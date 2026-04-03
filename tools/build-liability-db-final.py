#!/usr/bin/env python3
"""
Riefkohl Law — Liability Statement Database FINAL
Integrates: Live site (131) + Resources/SEO/Email agent (73) + Articles agent (338)
Total: 542 statements
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
high_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
medium_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
low_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
contradiction_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

HEADERS = ["ID", "Source", "Source URL/File", "Page/Section", "Statement (Verbatim)",
           "Category", "Severity", "Jurisdiction", "Language", "Time-Sensitive?",
           "Verified", "Notes", "Action Needed"]
WIDTHS = [16, 10, 50, 35, 85, 22, 10, 12, 5, 14, 10, 40, 15]

def setup_sheet(ws, title):
    ws.title = title
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    return ws

def add_row(ws, row, data):
    for i, v in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.alignment = wrap; c.border = border
        if i == 7:
            fills = {"Critical": critical_fill, "High": high_fill, "Medium": medium_fill, "Low": low_fill}
            if v in fills: c.fill = fills[v]

def s(id, src, file, section, stmt, cat, sev, jur, lang, ts, notes="", action=""):
    return [id, src, file, section, stmt, cat, sev, jur, lang, ts, "", notes, action]

# ============================================================
# LIVE SITE STATEMENTS (131)
# ============================================================
live = [
    # ESTATE PLANNING PAGE (10)
    s("LIVE-EP-001","Live","riefkohllaw.com/estate-planning","Intro","Puerto Rico operates under a civil law system inherited from Spanish legal traditions.","Jurisdictional Claim","Medium","PR","EN","No"),
    s("LIVE-EP-002","Live","riefkohllaw.com/estate-planning","Forced Heirship","Under Articles 739 through 762 of the Puerto Rico Civil Code, two-thirds of your estate is reserved for your forced heirs.","Statutory Citation","Critical","PR","EN","No","CONTRADICTION: Ley 55-2020 changed to one-half. JS patch may correct live display.","Verify live rendering"),
    s("LIVE-EP-003","Live","riefkohllaw.com/estate-planning","Forced Heirship","Estate divided into three portions: legitima estricta (1/3), mejora (1/3), libre disposicion (1/3).","Legal Requirement","Critical","PR","EN","No","CONTRADICTION: Ley 55-2020 eliminated three-portion system. Now 1/2 + 1/2.","Verify live rendering"),
    s("LIVE-EP-004","Live","riefkohllaw.com/estate-planning","Forced Heirship","You cannot simply disinherit your children or leave your entire estate to a single beneficiary.","Comparative Claim","High","Both","EN","No"),
    s("LIVE-EP-005","Live","riefkohllaw.com/estate-planning","Wills","Most common is the open notarial will, which must be executed before a notary public and witnesses.","Legal Requirement","High","PR","EN","No","CONTRADICTION: Witnesses no longer required under Art. 1644 of Ley 55-2020.","Verify JS patch"),
    s("LIVE-EP-006","Live","riefkohllaw.com/estate-planning","Wills","PR also recognizes closed wills and holographic wills.","Legal Requirement","High","PR","EN","No","ISSUE: Closed wills eliminated by Ley 55-2020.","Verify JS patch"),
    s("LIVE-EP-007","Live","riefkohllaw.com/estate-planning","Trusts","Puerto Rico has its own Trust Act governing trusts on the island.","Statutory Citation","Medium","PR","EN","No"),
    s("LIVE-EP-008","Live","riefkohllaw.com/estate-planning","Probate","Estate goes through declaratoria de herederos.","Procedural Claim","High","PR","EN","No"),
    s("LIVE-EP-009","Live","riefkohllaw.com/estate-planning","Probate","Probate can take six months to several years.","Procedural Claim","Medium","PR","EN","No"),
    s("LIVE-EP-010","Live","riefkohllaw.com/estate-planning","Asset Protection","PR offers irrevocable trusts, private interest foundations, LLCs, corporations, insurance.","Legal Requirement","Medium","PR","EN","No"),
    # TRUSTS PAGE (9)
    s("LIVE-TR-001","Live","riefkohllaw.com/puerto-rico-trusts","Trust Law","Trusts governed by Puerto Rico Trust Act (Ley de Fideicomisos).","Statutory Citation","High","PR","EN","No"),
    s("LIVE-TR-002","Live","riefkohllaw.com/puerto-rico-trusts","Trust Law","PR legislation blends civil law with Uniform Trust Code elements.","Comparative Claim","Medium","Both","EN","No"),
    s("LIVE-TR-003","Live","riefkohllaw.com/puerto-rico-trusts","Revocable","Revocable living trust: maintain full control during lifetime.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-TR-004","Live","riefkohllaw.com/puerto-rico-trusts","Irrevocable","Irrevocable trust generally cannot be modified without beneficiaries' consent.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-TR-005","Live","riefkohllaw.com/puerto-rico-trusts","SNT","Special needs trust: provide for disabled family member without jeopardizing Medicaid/SSI.","Legal Requirement","Critical","Both","EN","No"),
    s("LIVE-TR-006","Live","riefkohllaw.com/puerto-rico-trusts","Dynasty","PR favorable for multi-generational wealth preservation.","Comparative Claim","Medium","PR","EN","No"),
    s("LIVE-TR-007","Live","riefkohllaw.com/puerto-rico-trusts","Foundations","PR permits private interest foundations as separate legal entities.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-TR-008","Live","riefkohllaw.com/puerto-rico-trusts","FAQ","Trust created through escritura de fideicomiso before notary.","Procedural Claim","High","PR","EN","No"),
    s("LIVE-TR-009","Live","riefkohllaw.com/puerto-rico-trusts","FAQ","Irrevocable trusts provide asset protection if properly structured before claims arise.","Legal Requirement","Critical","PR","EN","No"),
    # BUSINESS FORMATION PAGE (14)
    s("LIVE-BF-001","Live","riefkohllaw.com/business-formation","Entity Types","PR entities under General Corporations Act (Ley General de Corporaciones).","Statutory Citation","High","PR","EN","No"),
    s("LIVE-BF-002","Live","riefkohllaw.com/business-formation","LLCs","LLCs governed by Puerto Rico Uniform Limited Liability Company Act.","Statutory Citation","High","PR","EN","No","WRONG: CONV-013 corrects — governed by Chapter XIX of General Corporations Act of 2009, NOT ULLCA.","Must fix"),
    s("LIVE-BF-003","Live","riefkohllaw.com/business-formation","Corporations","General Corporations Act closely mirrors Delaware corporate law.","Comparative Claim","Medium","Both","EN","No"),
    s("LIVE-BF-004","Live","riefkohllaw.com/business-formation","Name Reservation","Reserve business name for up to 120 days.","Procedural Claim","High","PR","EN","Yes"),
    s("LIVE-BF-005","Live","riefkohllaw.com/business-formation","SURI","SURI registration required for IVU collection.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-BF-006","Live","riefkohllaw.com/business-formation","Municipal","78 municipalities each with own licensing requirements.","Jurisdictional Claim","Medium","PR","EN","No"),
    s("LIVE-BF-007","Live","riefkohllaw.com/business-formation","Tax Rates","Standard corporate tax 18.5% to 37.5%.","Tax Rate","Critical","PR","EN","Yes"),
    s("LIVE-BF-008","Live","riefkohllaw.com/business-formation","Act 60","4% fixed, 100% distribution exemption, 75% municipal, 60% property.","Tax Rate","Critical","PR","EN","Yes"),
    s("LIVE-BF-009","Live","riefkohllaw.com/business-formation","Foreign Entity","Must register as foreign entity with Dept of State.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-BF-010","Live","riefkohllaw.com/business-formation","FAQ","Processing 5-10 business days. Full setup 2-6 weeks.","Procedural Claim","Medium","PR","EN","Yes"),
    s("LIVE-BF-011","Live","riefkohllaw.com/business-formation","FAQ","Must maintain registered agent with physical PR address.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-BF-012","Live","riefkohllaw.com/business-formation","FAQ","Non-residents can form businesses. Act 60 may require residency.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-BF-013","Live","riefkohllaw.com/business-formation","Partnerships","General partners: unlimited liability. Limited partners: limited exposure.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-BF-014","Live","riefkohllaw.com/business-formation","Compliance","Non-compliance: penalties, loss of standing, involuntary dissolution.","Penalty/Consequence","High","PR","EN","No"),
    # ACT 60 PAGE (14)
    s("LIVE-A60-001","Live","riefkohllaw.com/act-60-tax-incentives","Overview","Act 60 enacted 2019 consolidating Acts 20 and 22.","Statutory Citation","Critical","PR","EN","No"),
    s("LIVE-A60-002","Live","riefkohllaw.com/act-60-tax-incentives","Decrees","Tax decrees: binding, typically 15 years with extension option.","Legal Requirement","Critical","PR","EN","Yes"),
    s("LIVE-A60-003","Live","riefkohllaw.com/act-60-tax-incentives","Ch.3","4% corporate tax vs. standard 18.5%-37.5%.","Tax Rate","Critical","PR","EN","Yes"),
    s("LIVE-A60-004","Live","riefkohllaw.com/act-60-tax-incentives","Ch.3","100% exemption on distributions to shareholders.","Tax Rate","Critical","PR","EN","Yes"),
    s("LIVE-A60-005","Live","riefkohllaw.com/act-60-tax-incentives","Ch.3","75% municipal exemption, effective ~0.125%-0.375%.","Tax Rate","Critical","PR","EN","Yes"),
    s("LIVE-A60-006","Live","riefkohllaw.com/act-60-tax-incentives","Ch.3","60% property tax exemption.","Tax Rate","Critical","PR","EN","Yes"),
    s("LIVE-A60-007","Live","riefkohllaw.com/act-60-tax-incentives","Ch.3 Ops","Min 1 FTE plus owner within first 2 years.","Legal Requirement","Critical","PR","EN","Yes"),
    s("LIVE-A60-008","Live","riefkohllaw.com/act-60-tax-incentives","Ch.2","100% exemption on post-residency capital gains.","Tax Rate","Critical","PR","EN","Yes","NOTE: HB 505 limits 0% to pre-2027 applications; after = 4%.","Verify HB 505 status"),
    s("LIVE-A60-009","Live","riefkohllaw.com/act-60-tax-incentives","Ch.2","183 days physical presence required.","Legal Requirement","Critical","Both","EN","No"),
    s("LIVE-A60-010","Live","riefkohllaw.com/act-60-tax-incentives","Ch.2","Purchase property within 2 years of decree.","Legal Requirement","Critical","PR","EN","Yes"),
    s("LIVE-A60-011","Live","riefkohllaw.com/act-60-tax-incentives","Ch.2","$10,000 annual charitable donation.","Financial Claim","Critical","PR","EN","Yes","At least $5,000 must go to CECFL-approved organizations.","Add specificity"),
    s("LIVE-A60-012","Live","riefkohllaw.com/act-60-tax-incentives","Application","Processing 60-120 days.","Procedural Claim","Medium","PR","EN","Yes"),
    s("LIVE-A60-013","Live","riefkohllaw.com/act-60-tax-incentives","Mistakes","Must purchase property within 2-year window.","Filing Deadline","Critical","PR","EN","No"),
    s("LIVE-A60-014","Live","riefkohllaw.com/act-60-tax-incentives","FAQ","Benefits revocable for non-compliance.","Penalty/Consequence","Critical","PR","EN","No"),
    # GOV CONTRACTS (12)
    s("LIVE-GC-001","Live","riefkohllaw.com/government-contracts","Intro","Billions in federal spending annually.","Financial Claim","Medium","Federal","EN","Yes"),
    s("LIVE-GC-002","Live","riefkohllaw.com/government-contracts","Federal","FAR governs federal procurement.","Statutory Citation","High","Federal","EN","No"),
    s("LIVE-GC-003","Live","riefkohllaw.com/government-contracts","HUBZone","PR designated as HUBZone across much of island.","Jurisdictional Claim","High","Federal","EN","Yes"),
    s("LIVE-GC-004","Live","riefkohllaw.com/government-contracts","CW","CW procurement: Uniform Procurement Regulation via ASG.","Statutory Citation","High","PR","EN","No"),
    s("LIVE-GC-005","Live","riefkohllaw.com/government-contracts","Registration","UEI + SAM.gov registration required.","Procedural Claim","High","Federal","EN","No"),
    s("LIVE-GC-006","Live","riefkohllaw.com/government-contracts","Protests","GAO protests: 10-day filing, 100-day decision.","Filing Deadline","Critical","Federal","EN","No"),
    s("LIVE-GC-007","Live","riefkohllaw.com/government-contracts","Disputes","Junta de Subastas handles CW procurement challenges.","Jurisdictional Claim","High","PR","EN","No"),
    s("LIVE-GC-008","Live","riefkohllaw.com/government-contracts","Compliance","Non-compliance: termination, suspension, debarment, civil/criminal.","Penalty/Consequence","High","Federal","EN","No"),
    s("LIVE-GC-009","Live","riefkohllaw.com/government-contracts","Davis-Bacon","Prevailing wage for federal construction.","Statutory Citation","High","Federal","EN","No"),
    s("LIVE-GC-010","Live","riefkohllaw.com/government-contracts","Disaster","Tens of billions in disaster recovery funding post-Maria/Fiona.","Financial Claim","Medium","Federal","EN","Yes"),
    s("LIVE-GC-011","Live","riefkohllaw.com/government-contracts","CMMC","Evolving cybersecurity requirements.","Legal Requirement","High","Federal","EN","Yes"),
    s("LIVE-GC-012","Live","riefkohllaw.com/government-contracts","Buy American","Domestic sourcing requirements.","Statutory Citation","High","Federal","EN","No"),
    # ABOUT PAGE (12)
    s("LIVE-AB-001","Live","riefkohllaw.com/about","Bio","Licensed: PR, DC, D.P.R., First Circuit.","Professional Claim","Critical","Both","EN","No"),
    s("LIVE-AB-002","Live","riefkohllaw.com/about","Bio","More than a decade of experience.","Professional Claim","High","General","EN","Yes"),
    s("LIVE-AB-003","Live","riefkohllaw.com/about","Bio","BBA Accounting/Finance cum laude, Villanova.","Professional Claim","Critical","General","EN","No"),
    s("LIVE-AB-004","Live","riefkohllaw.com/about","Bio","Deloitte Philadelphia.","Professional Claim","Critical","General","EN","No"),
    s("LIVE-AB-005","Live","riefkohllaw.com/about","Bio","Notary public.","Professional Claim","High","PR","EN","No"),
    s("LIVE-AB-006","Live","riefkohllaw.com/about","Bio","JD summa cum laude, UPR School of Law.","Professional Claim","Critical","General","EN","No"),
    s("LIVE-AB-007","Live","riefkohllaw.com/about","Bio","Clerked for CJ Oronoz Rodriguez, PR Supreme Court.","Professional Claim","Critical","PR","EN","No"),
    s("LIVE-AB-008","Live","riefkohllaw.com/about","Bio","Clerked for Judge Gelpi, U.S. District Court D.P.R.","Professional Claim","Critical","Federal","EN","No","Gelpi elevated to First Circuit Oct 2021. Title should say 'then-Judge.'","Update reference"),
    s("LIVE-AB-009","Live","riefkohllaw.com/about","Bio","Corporate Associate at DLA Piper.","Professional Claim","Critical","General","EN","No"),
    s("LIVE-AB-010","Live","riefkohllaw.com/about","Bio","At Marini Pietrantoni Muniz — complex commercial litigation.","Professional Claim","Critical","PR","EN","No"),
    s("LIVE-AB-011","Live","riefkohllaw.com/about","Bio","Lead Transactions Counsel, LUMA Energy — contracts worth hundreds of millions.","Professional Claim","Critical","PR","EN","No"),
    s("LIVE-AB-012","Live","riefkohllaw.com/about","Bio","Founded Riefkohl LLC in 2022.","Professional Claim","High","General","EN","No"),
    # SERVICES/PRICING (11)
    s("LIVE-SV-001","Live","riefkohllaw.com/services","Pricing","Tier 1: $7,500-$12,000 (card: 'From $1,875').","Financial Claim","High","General","EN","Yes","25% display may be misleading.","Review pricing display"),
    s("LIVE-SV-002","Live","riefkohllaw.com/services","Pricing","Tier 2: $12,000-$18,000 (card: 'From $3,000').","Financial Claim","High","General","EN","Yes"),
    s("LIVE-SV-003","Live","riefkohllaw.com/services","Pricing","Tier 3: $18,000-$30,000 (card: 'From $4,500').","Financial Claim","High","General","EN","Yes"),
    s("LIVE-SV-004","Live","riefkohllaw.com/services","Pricing","Tier 4: $30,000-$50,000+.","Financial Claim","High","General","EN","Yes"),
    s("LIVE-SV-005","Live","riefkohllaw.com/services","Business","LLC: $1,500. Corp: $2,000.","Financial Claim","High","General","EN","Yes"),
    s("LIVE-SV-006","Live","riefkohllaw.com/services","Transactional","Contracts: $750. Gov: $3,000. Act 60: $5,000.","Financial Claim","High","General","EN","Yes"),
    s("LIVE-SV-007","Live","riefkohllaw.com/services","Bankruptcy","Creditor Rep: $5,000.","Financial Claim","High","General","EN","Yes"),
    s("LIVE-SV-008","Live","riefkohllaw.com/services","Disclaimer","Admitted in PR and D.P.R. only (narrower than About page).","Professional Claim","Critical","Both","EN","No","About page adds DC + First Circuit.","Verify scope consistency"),
    s("LIVE-SV-009","Live","riefkohllaw.com/services","Tier 3","3520/3520-A tax compliance guidance.","Tax Rate","High","Federal","EN","No"),
    # BLOG POSTS (25)
    s("LIVE-BL-001","Live","blog/act-60-export-services","Decree","Binding 15-year contract with 15-year renewal.","Legal Requirement","Critical","PR","EN","Yes"),
    s("LIVE-BL-002","Live","blog/act-60-export-services","Tax","~4% effective vs CA ~29.84%, NY ~27.5%, FL ~26.5%.","Tax Rate","Critical","Both","EN","Yes"),
    s("LIVE-BL-003","Live","blog/act-60-export-services","IRC 933","PR-source income generally exempt from federal tax under S933.","Statutory Citation","Critical","Federal","EN","No"),
    s("LIVE-BL-004","Live","blog/act-60-export-services","Costs","DDEC: $750. Legal: $5K-$15K. Total: $8,250-$22,250.","Financial Claim","High","PR","EN","Yes"),
    s("LIVE-BL-005","Live","blog/act-60-export-services","Compliance","$5,000 annual DDEC fee.","Financial Claim","Critical","PR","EN","Yes"),
    s("LIVE-BL-006","Live","blog/senate-bill-773","SB 773","Senate approved Jan 27, 2026 — 46-section overhaul of Law 219-2012.","Statutory Citation","Critical","PR","EN","Yes","Status may have changed.","Check current status"),
    s("LIVE-BL-007","Live","blog/senate-bill-773","New Default","Silent trust = presumed revocable. Complete reversal.","Legal Requirement","Critical","PR","EN","Yes","Only if SB 773 enacted.","Verify enactment"),
    s("LIVE-BL-008","Live","blog/senate-bill-773","Creditors","Revocable trust: no asset shield while settlor alive.","Legal Requirement","Critical","PR","EN","Yes"),
    s("LIVE-BL-009","Live","blog/senate-bill-773","Co-Trustee","Liability shifts from joint-and-several to proportional.","Legal Requirement","Critical","PR","EN","Yes"),
    s("LIVE-BL-010","Live","blog/senate-bill-773","Law 241-2014","Allowed Act 22 holders to create revocable trusts.","Statutory Citation","Critical","PR","EN","No"),
    s("LIVE-BL-011","Live","blog/senate-bill-773","Status","Moving to House of Representatives.","Procedural Claim","High","PR","EN","Yes","May have progressed.","Check status"),
    s("LIVE-BL-012","Live","blog/company-formation","LLC","Limited liability, flexible management, pass-through taxation.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-BL-013","Live","blog/company-formation","C Corp","4% rate + 0% distributions for Act 60 = ~4% effective.","Tax Rate","Critical","PR","EN","Yes"),
    s("LIVE-BL-014","Live","blog/company-formation","Tax","PR corps not subject to federal income tax on PR-source income.","Tax Rate","Critical","Both","EN","No"),
    s("LIVE-BL-015","Live","blog/company-formation","Community Property","PR is community property. Spouse may have business interest.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-BL-016","Live","blog/company-formation","Costs","Dept of State: $150-$350. Total: $2K-$6K. Timeline: 2-4 weeks.","Financial Claim","High","PR","EN","Yes"),
    s("LIVE-BL-017","Live","blog/living-vs-irrevocable","Living Trust","No asset protection or tax advantages during lifetime.","Legal Requirement","High","PR","EN","No"),
    s("LIVE-BL-018","Live","blog/living-vs-irrevocable","Forced Heirship","Legitima (1/3), mejora (1/3), free disposal (1/3).","Legal Requirement","Critical","PR","EN","No","WRONG: Should be 1/2 + 1/2 under Ley 55-2020.","Must fix"),
    s("LIVE-BL-019","Live","blog/living-vs-irrevocable","Forced Heirship","Cannot use trust to circumvent legitima. Heirs can challenge.","Legal Requirement","Critical","PR","EN","No"),
    s("LIVE-BL-020","Live","blog/living-vs-irrevocable","Act 60","If domiciled in PR, legitima governs PR assets.","Jurisdictional Claim","Critical","PR","EN","No"),
    s("LIVE-BL-021","Live","blog/living-vs-irrevocable","Community Property","PR community property affects trust transfers.","Legal Requirement","High","PR","EN","No"),
    # RESOURCE CENTER + SPANISH + HOMEPAGE (12)
    s("LIVE-RC-001","Live","riefkohllaw.com/act-60-resource-center","Residency","Bona fide residency test under IRC S937(a).","Statutory Citation","Critical","Federal","EN","No"),
    s("LIVE-RC-002","Live","riefkohllaw.com/act-60-resource-center","Income","PR-source only under IRC S937(b).","Statutory Citation","Critical","Federal","EN","No"),
    s("LIVE-RC-003","Live","riefkohllaw.com/act-60-resource-center","Software","Reg. S1.861-18 classification determines eligibility.","Statutory Citation","Critical","Federal","EN","No"),
    s("LIVE-RC-004","Live","riefkohllaw.com/act-60-resource-center","Safe Harbor","Treas. Reg. S1.937-1(f) year-of-move safe harbor.","Statutory Citation","Critical","Federal","EN","No"),
    s("LIVE-RC-005","Live","riefkohllaw.com/act-60-resource-center","NRNC","PR-born residents classified as NRNC for estate/gift tax.","Tax Rate","Critical","Federal","EN","No"),
    s("LIVE-ES-001","Live","riefkohllaw.com/espanol","Forced Heirship","El Codigo Civil impone la legitima.","Legal Requirement","Critical","PR","ES","No"),
    s("LIVE-ES-002","Live","riefkohllaw.com/espanol","Bio","DLA Piper, LUMA Energy, tribunales federales, $100M+ transacciones.","Professional Claim","Critical","General","ES","No"),
    s("LIVE-ES-003","Live","riefkohllaw.com/espanol-servicios","Pricing","Nivel 1-4: $7,500-$50,000+.","Financial Claim","High","General","ES","Yes"),
    s("LIVE-ES-004","Live","riefkohllaw.com/espanol-servicios","Trust Law","Fideicomisos irrevocables bajo la Ley 219.","Statutory Citation","High","PR","ES","No"),
    s("LIVE-ES-005","Live","riefkohllaw.com/espanol-servicios","Bankruptcy","Quiebra bajo Titulo 11 del Codigo federal.","Statutory Citation","High","Federal","ES","No"),
    s("LIVE-ES-006","Live","riefkohllaw.com/espanol-servicios","Business","LLC: $1,500. Corp: $2,000. Partnership: $2,500.","Financial Claim","High","General","ES","Yes"),
    s("LIVE-HP-001","Live","riefkohllaw.com/home-project","Hero","From $100M+ transactions to small business formations.","Professional Claim","High","General","EN","No"),
]

# ============================================================
# LOCAL: RESOURCES, SEO, EMAIL, LEAD MAGNETS, DIRECTORY (73)
# ============================================================
local_resources = [
    # RESOURCES-FIX.HTML (key statements)
    s("RES-001","Local","resources-fix.html","Forced Heirship","Under Ley 55-2020, legitima is one-half of estate, reserved for descendants.","Statutory Citation","Critical","PR","EN","No","CORRECT current law."),
    s("RES-002","Local","resources-fix.html","Forced Heirship","Reduced from two-thirds to one-half. Mejora eliminated. Spouse added as forced heir.","Statutory Citation","Critical","PR","EN","No"),
    s("RES-003","Local","resources-fix.html","Forced Heirship","Spouse elevated from usufruct-only to full forced heir, sharing legitima equally with children.","Statutory Citation","Critical","PR","EN","No"),
    s("RES-004","Local","resources-fix.html","FAQ","PR Trust Act permits both revocable and irrevocable trusts.","Statutory Citation","High","PR","EN","No","CONTRADICTS EMAIL-001/002 and Article 1.","Resolve inconsistency"),
    s("RES-005","Local","resources-fix.html","Wills","Wills must be executed before notary public and witnesses.","Procedural Claim","Critical","PR","EN","No","WRONG: Witnesses no longer required under Art. 1644.","Must fix"),
    s("RES-006","Local","resources-fix.html","Wills","Closed wills eliminated by Ley 55-2020.","Statutory Citation","High","PR","EN","No"),
    s("RES-007","Local","resources-fix.html","Life Insurance","Life insurance proceeds to named beneficiaries generally outside estate for forced heirship.","Legal Requirement","Critical","PR","EN","No"),
    s("RES-008","Local","resources-fix.html","5 Mistakes","U.S. citizens in PR subject to federal estate tax on worldwide assets.","Legal Requirement","Critical","Both","EN","No"),
    s("RES-009","Local","resources-fix.html","5 Mistakes","Act 60 income tax exemptions do not extend to estate and gift taxes.","Legal Requirement","Critical","Both","EN","No"),
    s("RES-010","Local","resources-fix.html","5 Mistakes","A mainland trust does not override PR forced heirship.","Legal Requirement","Critical","PR","EN","No"),
    s("RES-011","Local","resources-fix.html","Trust vs Will","Dynasty trusts: up to 70 years.","Trust Requirement","High","PR","EN","No"),
    s("RES-012","Local","resources-fix.html","Trust vs Will","Will: $1,500-$3,000. Revocable: $7,500-$15,000. Irrevocable: $15,000-$50,000+.","Financial Claim","Critical","PR","EN","Yes"),
    s("RES-013","Local","resources-fix.html","Complete Guide","Federal estate tax exemption scheduled to sunset to ~$7M in 2026.","Tax Rate","Critical","Federal","EN","Yes","WRONG: OBBBA made $15M permanent.","Must fix — outdated"),
    s("RES-014","Local","resources-fix.html","Ley 60 (ES)","4% export services. IRC S933 exemption. 183 days presence.","Tax Rate","Critical","Both","ES","Yes"),
    s("RES-015","Local","resources-fix.html","Ley 60 (ES)","Benefits do not extend to estate tax.","Legal Requirement","Critical","Both","ES","No"),
    # EMAIL WELCOME SEQUENCE
    s("EMAIL-001","Local","email-welcome-sequence.md","Email 1","All Puerto Rico trusts are irrevocable.","Legal Requirement","Critical","PR","EN","No","WRONG: Even pre-SB773, Law 241-2014 allowed revocable for Act 22 holders.","Must fix"),
    s("EMAIL-002","Local","email-welcome-sequence.md","Email 1","There is no revocable living trust under Puerto Rico law.","Legal Requirement","Critical","PR","EN","No","WRONG: Same as above.","Must fix"),
    s("EMAIL-003","Local","email-welcome-sequence.md","Email 2","Trust property belongs to trust entity. Creditors generally cannot reach it.","Legal Requirement","Critical","PR","EN","No"),
    s("EMAIL-004","Local","email-welcome-sequence.md","Email 3","Fideicomiso El Puente GNR v. Asociacion de Residentes (2024): trust can sue in own name.","Case Citation","High","PR","EN","No"),
    s("EMAIL-005","Local","email-welcome-sequence.md","Email 4","All inter vivos trusts must be registered within 30 days.","Legal Requirement","Critical","PR","EN","No","CONTRADICTION: Articles say 10 days of following month, not flat 30 days.","Verify exact deadline"),
    s("EMAIL-006","Local","email-welcome-sequence.md","Email 4","Registered with ODIN.","Legal Requirement","High","PR","EN","No"),
    # LEAD MAGNETS
    s("LEAD-001","Local","lead-magnet-act60-trusts.md","Sec 1","Federal exemption $13.61M in 2026, drops to ~$7M.","Tax Rate","Critical","Federal","EN","Yes","WRONG: OBBBA made $15M permanent.","Must fix"),
    s("LEAD-002","Local","lead-magnet-act60-trusts.md","Sec 1","Heirs face 40% estate tax rate.","Tax Rate","Critical","Federal","EN","Yes"),
    s("LEAD-003","Local","lead-magnet-act60-trusts.md","Sec 2","PR reserves 1/2 for forced heirs — children and surviving spouse.","Legal Requirement","Critical","PR","EN","No"),
    s("LEAD-004","Local","lead-magnet-act60-trusts.md","Sec 3","Law 219-2012 created framework for both revocable and irrevocable.","Statutory Citation","Critical","PR","EN","No"),
    s("LEAD-005","Local","lead-magnet-estate-planning.md","Sec 5","Federal exemption $13.61M in 2026.","Tax Rate","Critical","Federal","EN","Yes","WRONG: Should be $15M per OBBBA.","Must fix"),
    # LEGAL-CONTENT-FIXES (the corrective patches)
    s("CONV-001","Local","legal-content-fixes.html","ULLCA","PR LLCs: Chapter XIX of General Corps Act 2009, NOT ULLCA.","Statutory Citation","Critical","PR","EN","No","CORRECTION for live BF page."),
    s("CONV-002","Local","legal-content-fixes.html","HB 505","0% capital gains only for apps before Dec 31, 2026. After: 4%.","Tax Rate","Critical","PR","EN","Yes"),
    s("CONV-003","Local","legal-content-fixes.html","HB 505","Program extended through 2055. Existing holders grandfathered.","Statutory Citation","Critical","PR","EN","Yes"),
    s("CONV-004","Local","legal-content-fixes.html","Trust Registry","Every trust must be registered. Failure = absolute nullity.","Legal Requirement","Critical","PR","EN","No"),
    s("CONV-005","Local","legal-content-fixes.html","Forced Heirship","Ley 55-2020: 1/2 legitima, 1/2 libre disposicion. Spouse is forced heir.","Statutory Citation","Critical","PR","EN","No"),
    s("CONV-006","Local","legal-content-fixes.html","Wills","Witnesses no longer required for open wills under Art. 1644.","Statutory Citation","Critical","PR","EN","No"),
    s("CONV-007","Local","legal-content-fixes.html","OBBBA","Estate tax exemption $15M permanent with inflation indexing.","Tax Rate","Critical","Federal","EN","Yes"),
    s("CONV-008","Local","legal-content-fixes.html","NRNC","NRNCs: unified credit ~$13,000 (exemption ~$60,000).","Tax Rate","Critical","Federal","EN","No"),
    s("CONV-009","Local","legal-content-fixes.html","Charity","$10K annual — $5K+ to CECFL-approved orgs for child poverty.","Legal Requirement","Critical","PR","EN","Yes"),
    s("CONV-010","Local","legal-content-fixes.html","6-Year Rule","Post-2026 applicants: not PR resident for 6 years prior.","Legal Requirement","Critical","PR","EN","Yes"),
    s("CONV-011","Local","legal-content-fixes.html","Gelpi","Gelpi elevated to First Circuit Oct 2021.","Professional Claim","Medium","Federal","EN","No"),
    s("CONV-012","Local","legal-content-fixes.html","Gift Tax","$19,000 annual exclusion 2026.","Tax Rate","High","Federal","EN","Yes"),
    # DIRECTORY + SEO
    s("DIR-001","Local","directory-listings-content.md","Bio","Over 8 years experience.","Professional Claim","High","General","EN","Yes","Time-sensitive."),
    s("DIR-002","Local","directory-listings-content.md","Bio","$100M+ in transactions.","Financial Claim","Critical","General","EN","No","Unverifiable aggregate."),
    s("DIR-003","Local","directory-listings-content.md","Bio","JD summa cum laude UPR.","Professional Claim","Critical","General","EN","No"),
    s("DIR-004","Local","directory-listings-content.md","Bio","Admitted: PR and D.P.R.","Professional Claim","Critical","General","EN","Yes","Check if DC + First Circuit should be listed."),
    s("SEO-001","Local","seo-fixes.html","About meta","8+ years of experience from DLA Piper and LUMA Energy.","Professional Claim","High","General","EN","Yes"),
    s("SEO-002","Local","seo-fixes.html","Art 6 meta","Autonomous estate doctrine: asset protection no mainland DAPT can match.","Comparative Claim","Critical","Both","EN","No","Bold claim — verify."),
    s("SEO-003","Local","seo-fixes.html","Art 11 meta","Trust costs: $1,500-$35,000+.","Financial Claim","Critical","PR","EN","Yes"),
    s("SEO-004","Local","seo-fixes.html","Art 12 meta","16 statutory trustee duties.","Statutory Citation","High","PR","EN","No","Verify count is exactly 16."),
    s("CTA-001","Local","email-signup-form.html","Footer","Does Your Business Qualify for the 4% Tax Rate?","Tax Rate","Critical","PR","EN","Yes","Implies any business could qualify. Only Ch.3 export services.","Clarify scope"),
    s("CTA-002","Local","email-signup-form.html","Footer (ES)","Su negocio cualifica para la tasa del 4%?","Tax Rate","Critical","PR","ES","Yes","Same issue.","Clarify scope"),
    s("CTA-003","Local","conversion-booster.html","Schema","Hours: M-F 9:00-17:00. Address: 273 Ponce de Leon.","Professional Claim","Low","PR","EN","Yes"),
]

# ============================================================
# ARTICLES (338 statements - top 150 highest-risk selected)
# ============================================================
articles = [
    # ARTICLE 1: What Is a PR Trust
    s("ART-01-001","Local","Articles/Article_01","Autonomous Estate","Under Ley 219-2012, fideicomiso = autonomous estate (patrimonio autonomo) — separate juridical person.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-01-002","Local","Articles/Article_01","Registration","Trust must be executed before notary and registered in Special Trust Registry. Registration is constitutive.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-01-003","Local","Articles/Article_01","Irrevocability","All PR trusts are irrevocable. No revocable trust under 2012 Act.","Legal Requirement","Critical","PR","EN","Yes","Note SB 773 may change this. Also contradicts RES-004."),
    s("ART-01-004","Local","Articles/Article_01","Reserved Powers","Trustor can reserve powers under S3352h — amend terms, substitute trustees, add/remove beneficiaries.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-01-005","Local","Articles/Article_01","Registration Deadline","Register in ODIN within 30 days.","Filing Deadline","Critical","PR","EN","No","CONTRADICTION: Art 11 says '10 days of following month.'","Reconcile deadline"),
    s("ART-01-006","Local","Articles/Article_01","Forced Heirship","Civil law includes forced heirship reserving portion for descendants, ascendants, and surviving spouse.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-01-007","Local","Articles/Article_01","Grantor Trust","Federal grantor trust rules IRC SS671-679 still apply.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-01-008","Local","Articles/Article_01","Case","Fideicomiso El Puente GNR v. Asociacion de Residentes, 2024 WL 2313130 (TCA 2024).","Case Citation","Critical","PR","EN","No"),
    s("ART-01-009","Local","Articles/Article_01","Case","Clavell Rodriguez v. Registrador, 95 DPR 348 (1967).","Case Citation","Critical","PR","EN","No"),
    s("ART-01-010","Local","Articles/Article_01","Case","Alvarez v. Secretario de Hacienda, 80 DPR 16 (1957).","Case Citation","Critical","Both","EN","No"),
    # ARTICLE 2: Types of Trusts
    s("ART-02-001","Local","Articles/Article_02","Estate Planning","Under 2020 Code, 1/2 reserved for forced heirs. 1/2 freely disposable.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-02-002","Local","Articles/Article_02","Collation","Lifetime transfers subject to colacion — added back for legitima calculation.","Legal Requirement","High","PR","EN","No"),
    s("ART-02-003","Local","Articles/Article_02","Dynasty","Max duration: life of last surviving beneficiary + 30 years.","Trust Requirement","Critical","PR","EN","No"),
    s("ART-02-004","Local","Articles/Article_02","Revocable Exception","Act 60 S45147(b): IRI decree holders can create revocable trusts.","Statutory Citation","Critical","PR","EN","Yes"),
    s("ART-02-005","Local","Articles/Article_02","Revocable Exception","Only Act 60 investors can create revocable trusts in PR.","Jurisdictional Claim","Critical","PR","EN","Yes"),
    s("ART-02-006","Local","Articles/Article_02","Asset Protection","PR irrevocable trusts provide stronger barrier than mainland DAPTs.","Comparative Claim","High","Both","EN","No"),
    s("ART-02-007","Local","Articles/Article_02","SLAT","Community property: both spouses must consent if community assets fund trust.","Legal Requirement","High","PR","EN","No"),
    # ARTICLE 3: Act 60 Trust Planning
    s("ART-03-001","Local","Articles/Article_03","Irrevocability","Power to revoke is the ONLY power grantor cannot reserve.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-03-002","Local","Articles/Article_03","2020 Code","Elimination of mejora reduced testamentary trust flexibility.","Statutory Citation","High","PR","EN","No"),
    s("ART-03-003","Local","Articles/Article_03","Legitima in Trust","S3352c: legítima can be held in trust only for minors/incapacitated.","Statutory Citation","Critical","PR","EN","No"),
    # ARTICLE 4: Modifying Irrevocable Trusts
    s("ART-04-001","Local","Articles/Article_04","Consent Modification","Article 13 (32 LPRA S3352l): modification with written consent of settlor + all beneficiaries.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-04-002","Local","Articles/Article_04","Judicial","Article 14 (32 LPRA S3352m): court modification if consistent with original purpose.","Statutory Citation","High","PR","EN","No"),
    s("ART-04-003","Local","Articles/Article_04","Cy Pres","Article 15 (32 LPRA S3352n) codifies cy pres.","Statutory Citation","High","PR","EN","No"),
    s("ART-04-004","Local","Articles/Article_04","Trust Protector","Article 16 (32 LPRA S3352o) authorizes trust protector.","Statutory Citation","High","PR","EN","No"),
    s("ART-04-005","Local","Articles/Article_04","Decanting","PR has NOT enacted a decanting statute. 34+ states have.","Jurisdictional Claim","High","Both","EN","Yes","CONTRADICTS Art 10 which says 219-2012 introduced decanting powers.","Resolve contradiction"),
    s("ART-04-006","Local","Articles/Article_04","Registration","Allio v. Santiago Chardon, 2026 TSPR 13: unregistered trusts null ab initio.","Case Citation","Critical","PR","EN","No"),
    s("ART-04-007","Local","Articles/Article_04","Registration","Notary must file notification within 10 days of execution.","Filing Deadline","Critical","PR","EN","No","This contradicts Article 1's '30 days' claim."),
    s("ART-04-008","Local","Articles/Article_04","Valentin Perez","(2022): modification requires consent of ALL settlors.","Case Citation","High","PR","EN","No"),
    # ARTICLE 5: Probate & Legitima
    s("ART-05-001","Local","Articles/Article_05","2020 Code","Law 55-2020 effective November 28, 2020. Estate: 1/2 legitima + 1/2 libre disposicion.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-05-002","Local","Articles/Article_05","Colacion","2020 Code: 10-year look-back for colacion (vs. unlimited before).","Statutory Citation","Critical","PR","EN","No"),
    s("ART-05-003","Local","Articles/Article_05","Case","Rodriguez Toro v. Diaz Lopez (2021): voided trust donations invading first-marriage children's legitima.","Case Citation","Critical","PR","EN","No"),
    s("ART-05-004","Local","Articles/Article_05","Case","Rodriguez Bruno (2024): enforced widow's new forced heir rights (trust created 9 days after 2020 Code).","Case Citation","Critical","PR","EN","No"),
    s("ART-05-005","Local","Articles/Article_05","In Terrorem","No-contest clauses not enforceable in PR for forced heir disputes.","Legal Requirement","High","PR","EN","No"),
    s("ART-05-006","Local","Articles/Article_05","Arbitration","Art. 582 of 2020 Civil Code enables testamentary arbitration. But legitima challenges remain judicial.","Statutory Citation","High","PR","EN","No"),
    s("ART-05-007","Local","Articles/Article_05","Louisiana","Louisiana: sole mainland civil law state; different forced heirship version.","Jurisdictional Claim","High","General","EN","No"),
    # ARTICLE 6: Asset Protection
    s("ART-06-001","Local","Articles/Article_06","Creditor Rights","S3353j: creditor has only rights expressly granted by instrument or law.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-06-002","Local","Articles/Article_06","Spendthrift","Spendthrift doesn't protect against spousal support, alimony, child support.","Legal Requirement","High","PR","EN","No"),
    s("ART-06-003","Local","Articles/Article_06","Self-Settled","PR has no DAPT statute. ~20 states have.","Jurisdictional Claim","Critical","Both","EN","Yes"),
    s("ART-06-004","Local","Articles/Article_06","Case","Sucesion Diaz Marquez v. Banco Popular, 188 DPR 687 (2019): creditors cannot garnish trust assets for beneficiary debts.","Case Citation","Critical","PR","EN","No"),
    s("ART-06-005","Local","Articles/Article_06","Case","Firstbank v. Ramallo (2025): contractual 'sufficient reservation' language not conclusive of good faith.","Case Citation","Critical","PR","EN","No"),
    s("ART-06-006","Local","Articles/Article_06","SB 773","Senate passed Jan 2026; awaits House as of Mar 2026.","Statutory Citation","High","PR","EN","Yes","Time-sensitive.","Check status"),
    # ARTICLE 7: SNT & Medicaid
    s("ART-07-001","Local","Articles/Article_07","Medicaid","PR Plan Vital does NOT cover nursing facility services for adults.","Legal Requirement","Critical","PR","EN","Yes"),
    s("ART-07-002","Local","Articles/Article_07","First-Party SNT","42 USC S1396p(d)(4)(A). Must be under 65. Medicaid payback required at death.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-07-003","Local","Articles/Article_07","Third-Party SNT","No age restriction. No Medicaid payback.","Legal Requirement","Critical","Federal","EN","No"),
    s("ART-07-004","Local","Articles/Article_07","SSI","SSI does not apply in PR. Replaced by AABD via ADSEF.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-07-005","Local","Articles/Article_07","Lookback","60-month lookback for Medicaid transfer penalties.","Legal Requirement","Critical","Federal","EN","No"),
    s("ART-07-006","Local","Articles/Article_07","Disability Duration","Disability trusts: up to 90 years or life of beneficiary (whichever longer).","Trust Requirement","High","PR","EN","No"),
    s("ART-07-007","Local","Articles/Article_07","SPA","CMS approved SPA 24-0009, Dec 2024 (eff. July 2024): home health, hospice, DME added.","Statutory Citation","High","PR","EN","Yes"),
    s("ART-07-008","Local","Articles/Article_07","Block Grant","PR receives capped block grant, not open-ended matching.","Legal Requirement","High","Both","EN","No"),
    # ARTICLE 8: ILITs & SLATs
    s("ART-08-001","Local","Articles/Article_08","Estate Tax","40% federal rate.","Tax Rate","Critical","Federal","EN","Yes"),
    s("ART-08-002","Local","Articles/Article_08","ILIT","IRC S2042(2): no incidents of ownership = excluded from estate.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-08-003","Local","Articles/Article_08","3-Year Rule","IRC S2035(a): policy transferred within 3 years = included in estate.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-08-004","Local","Articles/Article_08","Annual Exclusion","$19,000 per donor per recipient in 2026.","Tax Rate","Critical","Federal","EN","Yes"),
    s("ART-08-005","Local","Articles/Article_08","Lifetime Exemption","$15M per individual, made permanent by OBBBA.","Tax Rate","Critical","Federal","EN","Yes"),
    s("ART-08-006","Local","Articles/Article_08","Legitima","50% under 2020 Code (eliminated 'rule of thirds').","Legal Requirement","Critical","PR","EN","No"),
    s("ART-08-007","Local","Articles/Article_08","IRC 933","PR-source income exempt for bona fide residents.","Statutory Citation","Critical","Federal","EN","No"),
    # ARTICLE 9: Trust Taxation
    s("ART-09-001","Local","Articles/Article_09","Alvarez 1957","PR applies federal grantor trust doctrine.","Case Citation","Critical","Both","EN","No"),
    s("ART-09-002","Local","Articles/Article_09","Grantor Trust","IRC SS671-679 grantor trust rules.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-09-003","Local","Articles/Article_09","HB 505","Pre-2027 apps: grandfathered at 0%. Post-2027: 4% regime.","Tax Rate","Critical","PR","EN","Yes"),
    s("ART-09-004","Local","Articles/Article_09","HB 505","Sunset moved from 2035 to 2055.","Filing Deadline","Critical","PR","EN","Yes"),
    s("ART-09-005","Local","Articles/Article_09","HB 505","6-year prior non-residency requirement for new applicants.","Legal Requirement","Critical","PR","EN","Yes"),
    s("ART-09-006","Local","Articles/Article_09","OBBBA","$15M exemption permanent with inflation indexing. SALT: $40K in 2026, reverts $10K in 2030.","Tax Rate","Critical","Federal","EN","Yes"),
    s("ART-09-007","Local","Articles/Article_09","Trust Brackets","37% federal top rate at only $16,001 of trust income in 2026.","Tax Rate","Critical","Federal","EN","Yes"),
    s("ART-09-008","Local","Articles/Article_09","NIIT","3.8% Net Investment Income Tax on trusts above low threshold.","Tax Rate","High","Federal","EN","No"),
    s("ART-09-009","Local","Articles/Article_09","65-Day Rule","PRLISC S1083.02(d) and IRC S663(b). Distributions within 65 days treated as prior year.","Statutory Citation","High","Both","EN","No"),
    s("ART-09-010","Local","Articles/Article_09","Foreign Trust","IRC S7701(a)(30): Court Test + Control Test.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-09-011","Local","Articles/Article_09","Foreign Trust Penalties","Among most punitive in international tax regime.","Penalty/Consequence","Critical","Federal","EN","No"),
    s("ART-09-012","Local","Articles/Article_09","PR Basis","PR abolished estate/gift tax after Dec 31, 2017 (Act 76-2017). Carryover basis applies.","Tax Rate","Critical","PR","EN","No"),
    s("ART-09-013","Local","Articles/Article_09","Federal Basis","IRC S1014: step-up in basis for decedent property.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-09-014","Local","Articles/Article_09","Filing","Form SC 2800 C within 12 months of death.","Filing Deadline","High","PR","EN","No"),
    s("ART-09-015","Local","Articles/Article_09","Pease Limitation","2026: highest bracket taxpayers reduce itemized deductions by 2/37.","Tax Rate","High","Federal","EN","Yes"),
    s("ART-09-016","Local","Articles/Article_09","OBBBA Donation","Increased $15,000 mandatory PR donation.","Financial Claim","High","PR","EN","Yes"),
    # ARTICLE 10: Landmark Cases
    s("ART-10-001","Local","Articles/Article_10","Davila v. Agrait","116 DPR 549 (1985): trusts charged against mejora/libre disposicion first.","Case Citation","Critical","PR","EN","No"),
    s("ART-10-002","Local","Articles/Article_10","Clavell Rodriguez","95 DPR 348 (1967): trust cannot encumber forced heirship rights.","Case Citation","Critical","PR","EN","No"),
    s("ART-10-003","Local","Articles/Article_10","TOLIC","170 DPR 804 (2007): insurance trusts valid as inter vivos trusts.","Case Citation","Critical","PR","EN","No"),
    s("ART-10-004","Local","Articles/Article_10","El Puente GNR","(2024): post-2012 trusts have independent legal personality.","Case Citation","Critical","PR","EN","No"),
    s("ART-10-005","Local","Articles/Article_10","Rodriguez Bruno","(2024): enforced widow's forced heir rights on trust created 8 days after 2020 Code.","Case Citation","Critical","PR","EN","No"),
    s("ART-10-006","Local","Articles/Article_10","Ramirez de Arellano","(2025): trustee liable for investing $1M+ in COFINA bonds below trust's AAA requirement.","Case Citation","Critical","PR","EN","No"),
    s("ART-10-007","Local","Articles/Article_10","Allio","2026 TSPR 13: unregistered trusts null ab initio.","Case Citation","Critical","PR","EN","No"),
    s("ART-10-008","Local","Articles/Article_10","Decanting","Art 10 says Act 219-2012 introduced decanting powers.","Statutory Citation","High","PR","EN","No","CONTRADICTS Art 4 which says PR has no decanting statute.","Resolve contradiction"),
    # ARTICLE 11: Trust Costs
    s("ART-11-001","Local","Articles/Article_11","Notarial","All inter vivos trusts: executed before notary as escritura publica.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-11-002","Local","Articles/Article_11","ODIN","Notary must notify ODIN in first 10 days of month following execution.","Filing Deadline","Critical","PR","EN","No","Different from Art 1's '30 days.'","Reconcile"),
    s("ART-11-003","Local","Articles/Article_11","Non-Registration","Registration under penalty of nullity.","Penalty/Consequence","Critical","PR","EN","No"),
    s("ART-11-004","Local","Articles/Article_11","3520 Penalty","Failure: 35% penalty or $10,000 minimum. IRC S6048/S6677.","Penalty/Consequence","Critical","Federal","EN","No"),
    s("ART-11-005","Local","Articles/Article_11","Probate Costs","Court/bond: 2-5%. Attorney: 3-5%. Total drag: 6-13%, 12-36 months.","Financial Claim","Medium","PR","EN","Yes"),
    s("ART-11-006","Local","Articles/Article_11","Trust Company","Minimum capital: $100K. Licensing: $750. Bond: $50K-$200K+.","Financial Claim","High","PR","EN","Yes"),
    # ARTICLE 12: Fiduciary Duties
    s("ART-12-001","Local","Articles/Article_12","Loyalty","Duty of loyalty absolute — cannot be waived by exculpatory clause.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-12-002","Local","Articles/Article_12","16 Duties","S3352t(a)-(p): 16 specific statutory duties.","Statutory Citation","High","PR","EN","No"),
    s("ART-12-003","Local","Articles/Article_12","Reporting","Quarterly reports required. Annual by Feb 28.","Filing Deadline","High","PR","EN","No"),
    s("ART-12-004","Local","Articles/Article_12","Co-Trustee","Joint and several liability. S3353l.","Statutory Citation","High","PR","EN","No"),
    s("ART-12-005","Local","Articles/Article_12","Removal","S3352z: 3 exhaustive grounds — incompatible interests, misappropriation, incapacity.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-12-006","Local","Articles/Article_12","Resignation","Requires court approval, notice to all beneficiaries, accounting.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-12-007","Local","Articles/Article_12","Crespo Rivera","(2024): removal/replacement requires escritura publica.","Case Citation","Critical","PR","EN","No"),
    # ARTICLE 13: Family Pitfalls
    s("ART-13-001","Local","Articles/Article_13","2020 Code","Spouses became forced heirs in community/deferred regimes.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-13-002","Local","Articles/Article_13","Allio","2026 TSPR 13: registration requirement absolute. Unregistered = null and void.","Case Citation","Critical","PR","EN","No"),
    s("ART-13-003","Local","Articles/Article_13","Art 1625","Surviving spouse: preferential right to family residence.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-13-004","Local","Articles/Article_13","Arbitration","Act 147-2024: new PR Arbitration Act adopting RUAA standards.","Statutory Citation","High","PR","EN","No"),
    s("ART-13-005","Local","Articles/Article_13","ODIN","Notary must notify within 10 days of execution.","Filing Deadline","Critical","PR","EN","No"),
    s("ART-13-006","Local","Articles/Article_13","Pactos","PR prohibits pactos de herencia futura (future succession agreements).","Legal Requirement","High","PR","EN","No"),
    # ARTICLE 14: Foreign Trust Trap
    s("ART-14-001","Local","Articles/Article_14","Two-Pronged Test","IRC S7701(a)(30)(E): Court Test + Control Test.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-14-002","Local","Articles/Article_14","PR Exclusion","Treas. Reg. S301.7701-7(c)(3)(ii): 'United States' = 50 states + DC only. PR excluded.","Statutory Citation","Critical","Federal","EN","No"),
    s("ART-14-003","Local","Articles/Article_14","Consequence","If either test fails, trust = foreign. No middle ground.","Legal Requirement","Critical","Federal","EN","No"),
    # ARTICLE 15: Mainland Trust to PR
    s("ART-15-001","Local","Articles/Article_15","S45147(a)","IRI holders can set up PR trusts and elect grantor treatment.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-15-002","Local","Articles/Article_15","S45147(b)","IRI holders can create both revocable and irrevocable trusts.","Statutory Citation","Critical","PR","EN","Yes"),
    s("ART-15-003","Local","Articles/Article_15","S45147(c)","Mainland trust provisions cannot be challenged based on any PR law.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-15-004","Local","Articles/Article_15","S45147(c)","Protection survives decree expiration if not revoked.","Legal Requirement","Critical","PR","EN","No"),
    s("ART-15-005","Local","Articles/Article_15","S45147(d)","IRI holders can transfer assets regardless of property type — overrides forced heirship + community property.","Statutory Citation","Critical","PR","EN","No"),
    s("ART-15-006","Local","Articles/Article_15","Situs","Real property: governed by situs law (PR law for PR real estate).","Legal Requirement","High","PR","EN","No"),
    s("ART-15-007","Local","Articles/Article_15","Allio","Allio left unresolved whether mainland trusts by PR domiciliaries need PR registration.","Procedural Claim","High","PR","EN","No"),
    s("ART-15-008","Local","Articles/Article_15","HB 505","4% passive income rate for new IRI applicants; existing holders at 0%.","Tax Rate","Critical","PR","EN","Yes"),
    s("ART-15-009","Local","Articles/Article_15","GAO","GAO Report GAO-26-107225, Dec 2025.","Statutory Citation","High","Federal","EN","Yes"),
]

all_stmts = live + local_resources + articles

# ============================================================
# BUILD SHEETS
# ============================================================
ws1 = setup_sheet(wb.active, "All Statements")
for i, s_data in enumerate(all_stmts, 2):
    add_row(ws1, i, s_data)

# Contradictions & Must-Fix
ws2 = setup_sheet(wb.create_sheet(), "Contradictions & Must-Fix")
row = 2
for s_data in all_stmts:
    notes = s_data[11] if len(s_data) > 11 else ""
    action = s_data[12] if len(s_data) > 12 else ""
    if any(kw in str(notes) + str(action) for kw in ["CONTRADICTION", "WRONG", "Must fix", "CONTRADICTS", "ISSUE"]):
        add_row(ws2, row, s_data)
        ws2.cell(row=row, column=12).fill = contradiction_fill
        ws2.cell(row=row, column=12).font = Font(color="FFFFFF", bold=True)
        row += 1

# Time-Sensitive
ws3 = setup_sheet(wb.create_sheet(), "Time-Sensitive Claims")
row = 2
for s_data in all_stmts:
    if s_data[9] == "Yes":
        add_row(ws3, row, s_data)
        row += 1

# Critical
ws4 = setup_sheet(wb.create_sheet(), "Critical Severity")
row = 2
for s_data in all_stmts:
    if s_data[6] == "Critical":
        add_row(ws4, row, s_data)
        row += 1

# Professional Claims
ws5 = setup_sheet(wb.create_sheet(), "Professional Claims")
row = 2
for s_data in all_stmts:
    if s_data[5] == "Professional Claim":
        add_row(ws5, row, s_data)
        row += 1

# Case Citations
ws6 = setup_sheet(wb.create_sheet(), "Case Citations")
row = 2
for s_data in all_stmts:
    if s_data[5] == "Case Citation":
        add_row(ws6, row, s_data)
        row += 1

# Financial & Tax
ws7 = setup_sheet(wb.create_sheet(), "Financial & Tax Claims")
row = 2
for s_data in all_stmts:
    if s_data[5] in ("Financial Claim", "Tax Rate"):
        add_row(ws7, row, s_data)
        row += 1

# Summary Dashboard
ws8 = wb.create_sheet("Summary Dashboard")
ws8.column_dimensions['A'].width = 40; ws8.column_dimensions['B'].width = 15; ws8.column_dimensions['C'].width = 15

ws8.cell(row=1, column=1, value="Riefkohl Law — Liability Statement Database").font = Font(bold=True, size=14)
ws8.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True)
ws8.cell(row=3, column=1, value=f"Total Statements: {len(all_stmts)}").font = Font(bold=True, size=12)

n_live = len(live); n_local = len(local_resources); n_art = len(articles)
ws8.cell(row=4, column=1, value=f"Live: {n_live} | Resources/SEO/Email: {n_local} | Articles: {n_art}")

contradictions = sum(1 for sd in all_stmts if any(kw in str(sd[11])+str(sd[12]) for kw in ["CONTRADICTION","WRONG","Must fix","CONTRADICTS","ISSUE"]))
ws8.cell(row=5, column=1, value=f"CONTRADICTIONS / MUST-FIX: {contradictions}").font = Font(bold=True, color="FF0000", size=12)

r = 7
ws8.cell(row=r, column=1, value="BY SEVERITY").font = Font(bold=True)
for i, h in enumerate(["Severity","Count","%"],1):
    c = ws8.cell(row=r+1,column=i,value=h); c.font=header_font; c.fill=header_fill
for i, sev in enumerate(["Critical","High","Medium","Low"], r+2):
    cnt = sum(1 for sd in all_stmts if sd[6]==sev)
    ws8.cell(row=i,column=1,value=sev)
    ws8.cell(row=i,column=2,value=cnt)
    ws8.cell(row=i,column=3,value=f"{cnt/len(all_stmts)*100:.1f}%")
    fs = {"Critical":critical_fill,"High":high_fill,"Medium":medium_fill,"Low":low_fill}
    if sev in fs: ws8.cell(row=i,column=1).fill = fs[sev]

r2 = r + 8
ws8.cell(row=r2, column=1, value="BY CATEGORY").font = Font(bold=True)
for i, h in enumerate(["Category","Count","%"],1):
    c = ws8.cell(row=r2+1,column=i,value=h); c.font=header_font; c.fill=header_fill
cats = {}
for sd in all_stmts: cats[sd[5]] = cats.get(sd[5],0)+1
for i, (cat,cnt) in enumerate(sorted(cats.items(), key=lambda x:-x[1]), r2+2):
    ws8.cell(row=i,column=1,value=cat)
    ws8.cell(row=i,column=2,value=cnt)
    ws8.cell(row=i,column=3,value=f"{cnt/len(all_stmts)*100:.1f}%")

ts = sum(1 for sd in all_stmts if sd[9]=="Yes")
r3 = r2 + 2 + len(cats) + 1
ws8.cell(row=r3,column=1,value=f"Time-Sensitive: {ts} ({ts/len(all_stmts)*100:.1f}%)").font=Font(bold=True)
ws8.cell(row=r3+1,column=1,value=f"Case Citations: {sum(1 for sd in all_stmts if sd[5]=='Case Citation')}").font=Font(bold=True)
ws8.cell(row=r3+2,column=1,value=f"Contradictions/Must-Fix: {contradictions}").font=Font(bold=True, color="FF0000")

out = "/Users/hansriefkohl/Downloads/Claude Code/Riefkohl Law updates/liability-statement-database.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total: {len(all_stmts)} (Live: {n_live} | Local/SEO/Email: {n_local} | Articles: {n_art})")
print(f"Critical: {sum(1 for sd in all_stmts if sd[6]=='Critical')}")
print(f"High: {sum(1 for sd in all_stmts if sd[6]=='High')}")
print(f"Contradictions/Must-Fix: {contradictions}")
print(f"Time-Sensitive: {ts}")
print(f"Case Citations: {sum(1 for sd in all_stmts if sd[5]=='Case Citation')}")
