#!/usr/bin/env python3
"""
Riefkohl Law — Liability Statement Database Builder v2
Integrates: Live site (135) + Resources/Homepage/SEO/Email agent (147) + Critical contradictions
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# Styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
high_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
medium_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
low_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
contradiction_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

HEADERS = ["ID", "Source", "Source URL/File", "Page/Section", "Statement (Verbatim)",
           "Category", "Severity", "Jurisdiction", "Language", "Time-Sensitive?",
           "Verified", "Notes", "Action Needed"]
WIDTHS = [16, 10, 50, 35, 85, 22, 10, 12, 5, 14, 10, 40, 15]

def setup_sheet(ws, title):
    ws.title = title
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    return ws

def add_row(ws, row, data):
    for i, v in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.alignment = wrap; c.border = border
        if i == 7:
            fills = {"Critical": critical_fill, "High": high_fill, "Medium": medium_fill, "Low": low_fill}
            if v in fills: c.fill = fills[v]

# ============================================================
# ALL STATEMENTS
# ============================================================

# --- LIVE SITE (135 statements from Phase 1) ---
live = [
    # ESTATE PLANNING PAGE
    ["LIVE-EP-001","Live","riefkohllaw.com/estate-planning","Intro","Puerto Rico operates under a civil law system inherited from Spanish legal traditions, unlike the common law system used in most U.S. states.","Jurisdictional Claim","Medium","PR","EN","No","","",""],
    ["LIVE-EP-002","Live","riefkohllaw.com/estate-planning","Forced Heirship","Under Articles 739 through 762 of the Puerto Rico Civil Code, two-thirds of your estate is reserved for your forced heirs, typically your children or descendants.","Statutory Citation","Critical","PR","EN","No","","CONTRADICTION: legal-content-fixes.html corrects this to one-half under Ley 55-2020. Live page may show corrected version via JS patch.","Verify live rendering"],
    ["LIVE-EP-003","Live","riefkohllaw.com/estate-planning","Forced Heirship","Your estate is divided into three portions: the legitima estricta (one-third), the mejora (one-third), and the libre disposicion (one-third).","Legal Requirement","Critical","PR","EN","No","","CONTRADICTION: Ley 55-2020 eliminated the three-portion system. Now two portions: 1/2 legitima, 1/2 libre disposicion. Check if JS patch corrects this.","Verify live rendering"],
    ["LIVE-EP-004","Live","riefkohllaw.com/estate-planning","Forced Heirship","This means you cannot simply disinherit your children or leave your entire estate to a single beneficiary.","Comparative Claim","High","Both","EN","No","","",""],
    ["LIVE-EP-005","Live","riefkohllaw.com/estate-planning","Wills","The most common is the open notarial will (testamento abierto), which must be executed before a notary public and witnesses.","Legal Requirement","High","PR","EN","No","","CONTRADICTION: CONV-030 says witnesses no longer required under Art. 1644 of Ley 55-2020.","Verify if JS patch corrects"],
    ["LIVE-EP-006","Live","riefkohllaw.com/estate-planning","Wills","Puerto Rico also recognizes closed wills and holographic wills.","Legal Requirement","High","PR","EN","No","","ISSUE: Closed wills eliminated by Ley 55-2020 per CONV-029.","Verify if JS patch corrects"],
    ["LIVE-EP-007","Live","riefkohllaw.com/estate-planning","Trusts","Puerto Rico has its own Trust Act that governs the creation and administration of trusts.","Statutory Citation","Medium","PR","EN","No","","",""],
    ["LIVE-EP-008","Live","riefkohllaw.com/estate-planning","Probate","Estate typically goes through a probate process known as declaratoria de herederos.","Procedural Claim","High","PR","EN","No","","",""],
    ["LIVE-EP-009","Live","riefkohllaw.com/estate-planning","Probate","The probate process can take anywhere from six months to several years.","Procedural Claim","Medium","PR","EN","No","","",""],
    ["LIVE-EP-010","Live","riefkohllaw.com/estate-planning","Asset Protection","Puerto Rico offers asset protection tools including irrevocable trusts, private interest foundations, LLCs/corporations, and insurance.","Legal Requirement","Medium","PR","EN","No","","",""],

    # TRUSTS PAGE
    ["LIVE-TR-001","Live","riefkohllaw.com/puerto-rico-trusts","Trust Law","Puerto Rico trusts are governed by the Puerto Rico Trust Act (Ley de Fideicomisos).","Statutory Citation","High","PR","EN","No","","",""],
    ["LIVE-TR-002","Live","riefkohllaw.com/puerto-rico-trusts","Trust Law","PR's trust legislation blends civil law concepts with elements from the Uniform Trust Code, creating a hybrid system.","Comparative Claim","Medium","Both","EN","No","","",""],
    ["LIVE-TR-003","Live","riefkohllaw.com/puerto-rico-trusts","Revocable Trust","A revocable living trust allows you to maintain full control over your assets during your lifetime.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-TR-004","Live","riefkohllaw.com/puerto-rico-trusts","Irrevocable Trust","Once established, an irrevocable trust generally cannot be modified or revoked without the consent of the beneficiaries.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-TR-005","Live","riefkohllaw.com/puerto-rico-trusts","Special Needs","A special needs trust enables you to provide for a disabled family member without jeopardizing eligibility for Medicaid or SSI.","Legal Requirement","Critical","Both","EN","No","","",""],
    ["LIVE-TR-006","Live","riefkohllaw.com/puerto-rico-trusts","Dynasty Trusts","Puerto Rico's favorable trust laws make it an attractive jurisdiction for multi-generational wealth preservation.","Comparative Claim","Medium","PR","EN","No","","",""],
    ["LIVE-TR-007","Live","riefkohllaw.com/puerto-rico-trusts","Foundations","Puerto Rico law permits the creation of private interest foundations, which operate as separate legal entities.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-TR-008","Live","riefkohllaw.com/puerto-rico-trusts","FAQ","A trust is created through a written trust agreement (escritura de fideicomiso) executed before a notary public.","Procedural Claim","High","PR","EN","No","","",""],
    ["LIVE-TR-009","Live","riefkohllaw.com/puerto-rico-trusts","FAQ","Irrevocable trusts can provide significant asset protection from creditors, provided properly structured and funded before claims arise.","Legal Requirement","Critical","PR","EN","No","","",""],

    # BUSINESS FORMATION PAGE
    ["LIVE-BF-001","Live","riefkohllaw.com/business-formation","Entity Types","Puerto Rico recognizes several business entity types under its General Corporations Act (Ley General de Corporaciones).","Statutory Citation","High","PR","EN","No","","",""],
    ["LIVE-BF-002","Live","riefkohllaw.com/business-formation","LLCs","LLCs in Puerto Rico are governed by the Puerto Rico Uniform Limited Liability Company Act.","Statutory Citation","High","PR","EN","No","","CONTRADICTION: CONV-013 corrects this — PR LLCs governed by Chapter XIX of General Corporations Act of 2009, NOT the ULLCA.","WRONG — must fix"],
    ["LIVE-BF-003","Live","riefkohllaw.com/business-formation","Corporations","The island's General Corporations Act closely mirrors Delaware corporate law.","Comparative Claim","Medium","Both","EN","No","","",""],
    ["LIVE-BF-004","Live","riefkohllaw.com/business-formation","Name Reservation","We can reserve the business name for up to 120 days.","Procedural Claim","High","PR","EN","Yes","","",""],
    ["LIVE-BF-005","Live","riefkohllaw.com/business-formation","SURI","Registration with SURI is required for local tax compliance, including IVU collection.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BF-006","Live","riefkohllaw.com/business-formation","Municipal","Each of Puerto Rico's 78 municipalities has its own licensing requirements.","Jurisdictional Claim","Medium","PR","EN","No","","",""],
    ["LIVE-BF-007","Live","riefkohllaw.com/business-formation","Tax Rates","The standard corporate tax rate in Puerto Rico ranges from 18.5% to 37.5%.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-BF-008","Live","riefkohllaw.com/business-formation","Act 60","4% fixed income tax, 100% exemption on distributions, 75% municipal exemption, 60% property tax exemption.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-BF-009","Live","riefkohllaw.com/business-formation","Foreign Entity","Must register as a foreign entity with the Department of State to operate in PR.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BF-010","Live","riefkohllaw.com/business-formation","FAQ","Standard processing takes 5-10 business days. Full setup 2-6 weeks.","Procedural Claim","Medium","PR","EN","Yes","","",""],
    ["LIVE-BF-011","Live","riefkohllaw.com/business-formation","FAQ","All business entities must maintain a registered agent with a physical address on the island.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BF-012","Live","riefkohllaw.com/business-formation","FAQ","Non-residents can form businesses with no residency requirement. Act 60 incentives may require residency.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BF-013","Live","riefkohllaw.com/business-formation","Partnerships","In a general partnership, all partners share liability. Limited partnerships allow passive investors to limit exposure.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BF-014","Live","riefkohllaw.com/business-formation","Compliance","Non-compliance can result in penalties, loss of good standing, or involuntary dissolution.","Penalty/Consequence","High","PR","EN","No","","",""],

    # ACT 60 PAGE
    ["LIVE-A60-001","Live","riefkohllaw.com/act-60-tax-incentives","Act 60 Overview","Act 60 enacted in 2019 to consolidate previous laws including Act 20 and Act 22.","Statutory Citation","Critical","PR","EN","No","","",""],
    ["LIVE-A60-002","Live","riefkohllaw.com/act-60-tax-incentives","Decrees","Tax decrees are legally binding agreements, typically 15 years with option to extend.","Legal Requirement","Critical","PR","EN","Yes","","",""],
    ["LIVE-A60-003","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 3 Rates","4% fixed corporate income tax rate vs. standard 18.5%-37.5%.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-A60-004","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 3 Rates","100% tax exemption on distributions from earnings and profits to shareholders.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-A60-005","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 3 Rates","75% exemption on municipal license taxes, reducing effective rate to ~0.125%-0.375%.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-A60-006","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 3 Rates","60% exemption on personal and real property taxes.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-A60-007","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 3 Operations","Must hire at least one full-time employee in addition to owner within first two years.","Legal Requirement","Critical","PR","EN","Yes","","",""],
    ["LIVE-A60-008","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 2 Benefits","100% exemption on capital gains accrued after becoming bona fide PR resident.","Tax Rate","Critical","PR","EN","Yes","","NOTE: CONV-014 says 0% rate only for applications before Dec 31 2026 under HB 505. After that, 4%.","Verify current status"],
    ["LIVE-A60-009","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 2 Residency","Physically present in PR for at least 183 days per year.","Legal Requirement","Critical","Both","EN","No","","",""],
    ["LIVE-A60-010","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 2 Residency","Purchase residential property within two years of decree grant.","Legal Requirement","Critical","PR","EN","Yes","","",""],
    ["LIVE-A60-011","Live","riefkohllaw.com/act-60-tax-incentives","Ch. 2 Residency","Annual charitable donation of $10,000 to qualifying Puerto Rico nonprofits.","Financial Claim","Critical","PR","EN","Yes","","CONV-035/036: At least $5,000 must go to CECFL-approved orgs focused on eradicating child poverty.","Add CECFL specificity"],
    ["LIVE-A60-012","Live","riefkohllaw.com/act-60-tax-incentives","Application","Typical processing time 60-120 days.","Procedural Claim","Medium","PR","EN","Yes","","",""],
    ["LIVE-A60-013","Live","riefkohllaw.com/act-60-tax-incentives","Mistakes","Not purchasing property within required two-year window.","Filing Deadline","Critical","PR","EN","No","","",""],
    ["LIVE-A60-014","Live","riefkohllaw.com/act-60-tax-incentives","FAQ","Benefits can be revoked for non-compliance.","Penalty/Consequence","Critical","PR","EN","No","","",""],

    # GOVERNMENT CONTRACTS PAGE
    ["LIVE-GC-001","Live","riefkohllaw.com/government-contracts","Intro","Billions of dollars in federal spending flowing to the island annually.","Financial Claim","Medium","Federal","EN","Yes","","",""],
    ["LIVE-GC-002","Live","riefkohllaw.com/government-contracts","Federal","Federal procurement governed by the FAR.","Statutory Citation","High","Federal","EN","No","","",""],
    ["LIVE-GC-003","Live","riefkohllaw.com/government-contracts","HUBZone","PR's designation as a HUBZone across much of the island.","Jurisdictional Claim","High","Federal","EN","Yes","","",""],
    ["LIVE-GC-004","Live","riefkohllaw.com/government-contracts","Commonwealth","Commonwealth procurement governed by PR's Uniform Procurement Regulation, administered through ASG.","Statutory Citation","High","PR","EN","No","","",""],
    ["LIVE-GC-005","Live","riefkohllaw.com/government-contracts","Registration","Must obtain UEI and register in SAM.gov.","Procedural Claim","High","Federal","EN","No","","",""],
    ["LIVE-GC-006","Live","riefkohllaw.com/government-contracts","Bid Protests","GAO protests must be filed within 10 days. GAO issues decisions within 100 days.","Filing Deadline","Critical","Federal","EN","No","","",""],
    ["LIVE-GC-007","Live","riefkohllaw.com/government-contracts","Disputes","Junta de Subastas handles many Commonwealth procurement challenges.","Jurisdictional Claim","High","PR","EN","No","","",""],
    ["LIVE-GC-008","Live","riefkohllaw.com/government-contracts","Compliance","Non-compliance can result in termination, suspension, debarment, civil/criminal liability.","Penalty/Consequence","High","Federal","EN","No","","",""],
    ["LIVE-GC-009","Live","riefkohllaw.com/government-contracts","Davis-Bacon","Davis-Bacon Act — Prevailing wage requirements for federally funded construction.","Statutory Citation","High","Federal","EN","No","","",""],
    ["LIVE-GC-010","Live","riefkohllaw.com/government-contracts","Disaster Recovery","PR received tens of billions in federal disaster recovery funding following Maria and Fiona.","Financial Claim","Medium","Federal","EN","Yes","","",""],
    ["LIVE-GC-011","Live","riefkohllaw.com/government-contracts","CMMC","CMMC requirements for contractors handling controlled unclassified information.","Legal Requirement","High","Federal","EN","Yes","","",""],
    ["LIVE-GC-012","Live","riefkohllaw.com/government-contracts","Buy American","Buy American Act — Requirements for domestic sourcing.","Statutory Citation","High","Federal","EN","No","","",""],

    # ABOUT PAGE
    ["LIVE-AB-001","Live","riefkohllaw.com/about","Bio","Licensed in Puerto Rico, Washington D.C., U.S. District Court for D.P.R., and U.S. Court of Appeals for the First Circuit.","Professional Claim","Critical","Both","EN","No","","",""],
    ["LIVE-AB-002","Live","riefkohllaw.com/about","Bio","More than a decade of experience across accounting, corporate transactions, government contracting, and complex litigation.","Professional Claim","High","General","EN","Yes","","",""],
    ["LIVE-AB-003","Live","riefkohllaw.com/about","Bio","BBA in Accounting and Finance, cum laude, from Villanova University.","Professional Claim","Critical","General","EN","No","","",""],
    ["LIVE-AB-004","Live","riefkohllaw.com/about","Bio","Spent time at Deloitte in Philadelphia.","Professional Claim","Critical","General","EN","No","","",""],
    ["LIVE-AB-005","Live","riefkohllaw.com/about","Bio","Served as a notary public.","Professional Claim","High","PR","EN","No","","",""],
    ["LIVE-AB-006","Live","riefkohllaw.com/about","Bio","J.D. summa cum laude from University of Puerto Rico School of Law.","Professional Claim","Critical","General","EN","No","","",""],
    ["LIVE-AB-007","Live","riefkohllaw.com/about","Bio","Clerked for Chief Justice Maite D. Oronoz Rodriguez at Supreme Court of PR.","Professional Claim","Critical","PR","EN","No","","",""],
    ["LIVE-AB-008","Live","riefkohllaw.com/about","Bio","Clerked for Judge Gustavo A. Gelpi at U.S. District Court for D.P.R.","Professional Claim","Critical","Federal","EN","No","","CONV-037: Gelpi was elevated to First Circuit in Oct 2021. Title should say 'then-Judge' or note elevation.","Update title reference"],
    ["LIVE-AB-009","Live","riefkohllaw.com/about","Bio","Corporate Associate at DLA Piper — PPPs, cross-border transactions, government contracting.","Professional Claim","Critical","General","EN","No","","",""],
    ["LIVE-AB-010","Live","riefkohllaw.com/about","Bio","At Marini Pietrantoni Muniz — complex commercial litigation.","Professional Claim","Critical","PR","EN","No","","",""],
    ["LIVE-AB-011","Live","riefkohllaw.com/about","Bio","Lead Transactions Counsel for LUMA Energy — contracts worth hundreds of millions.","Professional Claim","Critical","PR","EN","No","","",""],
    ["LIVE-AB-012","Live","riefkohllaw.com/about","Bio","In 2022, Hans founded Riefkohl LLC.","Professional Claim","High","General","EN","No","","",""],

    # SERVICES/PRICING
    ["LIVE-SV-001","Live","riefkohllaw.com/services","Trust Pricing","Tier 1: $7,500-$12,000 (card shows 'From $1,875').","Financial Claim","High","General","EN","Yes","","ISSUE: 'From $1,875' is 25% of $7,500. Potentially misleading if visitors think $1,875 is the actual price.","Review pricing display"],
    ["LIVE-SV-002","Live","riefkohllaw.com/services","Trust Pricing","Tier 2: $12,000-$18,000 (card shows 'From $3,000').","Financial Claim","High","General","EN","Yes","","Same 25% display issue.","Review pricing display"],
    ["LIVE-SV-003","Live","riefkohllaw.com/services","Trust Pricing","Tier 3: $18,000-$30,000 (card shows 'From $4,500').","Financial Claim","High","General","EN","Yes","","Same 25% display issue.","Review pricing display"],
    ["LIVE-SV-004","Live","riefkohllaw.com/services","Trust Pricing","Tier 4: $30,000-$50,000+ (card shows 'From $7,500').","Financial Claim","High","General","EN","Yes","","Same 25% display issue.","Review pricing display"],
    ["LIVE-SV-005","Live","riefkohllaw.com/services","Business","LLC Formation Starting at $1,500.","Financial Claim","High","General","EN","Yes","","",""],
    ["LIVE-SV-006","Live","riefkohllaw.com/services","Business","Corporation Formation Starting at $2,000.","Financial Claim","High","General","EN","Yes","","",""],
    ["LIVE-SV-007","Live","riefkohllaw.com/services","Transactional","Contract Drafting & Review Starting at $750.","Financial Claim","High","General","EN","Yes","","",""],
    ["LIVE-SV-008","Live","riefkohllaw.com/services","Transactional","Government Contract Support Starting at $3,000.","Financial Claim","High","General","EN","Yes","","",""],
    ["LIVE-SV-009","Live","riefkohllaw.com/services","Transactional","Act 60 Services Starting at $5,000.","Financial Claim","High","General","EN","Yes","","",""],
    ["LIVE-SV-010","Live","riefkohllaw.com/services","Bankruptcy","Creditor Representation Starting at $5,000.","Financial Claim","High","General","EN","Yes","","",""],
    ["LIVE-SV-011","Live","riefkohllaw.com/services","Disclaimer","Admitted to practice in PR and before U.S. District Court for D.P.R.","Professional Claim","Critical","Both","EN","No","","About page says also D.C. and First Circuit. This disclaimer is narrower.","Verify scope matches"],

    # BLOG: ACT 60 TAX GUIDE
    ["LIVE-BL-A60-001","Live","blog/act-60-export-services-guide","Decree","Decree is binding contract, typically 15 years with 15-year renewal option.","Legal Requirement","Critical","PR","EN","Yes","","",""],
    ["LIVE-BL-A60-002","Live","blog/act-60-export-services-guide","Municipal Tax","75% exemption reduces effective rate to ~0.125%-0.375%.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-BL-A60-003","Live","blog/act-60-export-services-guide","Tax Comparison","PR Act 60 ~4% vs California ~29.84% vs New York ~27.5% vs Florida ~26.5%.","Tax Rate","Critical","Both","EN","Yes","","",""],
    ["LIVE-BL-A60-004","Live","blog/act-60-export-services-guide","Section 933","PR-source income under valid decree generally not subject to federal income tax under IRC S933.","Statutory Citation","Critical","Federal","EN","No","","",""],
    ["LIVE-BL-A60-005","Live","blog/act-60-export-services-guide","Costs","DDEC Filing: $750. Legal: $5,000-$15,000. CPA: $2,000-$5,000. Total: $8,250-$22,250.","Financial Claim","High","PR","EN","Yes","","",""],
    ["LIVE-BL-A60-006","Live","blog/act-60-export-services-guide","Compliance","$5,000 annual compliance fee to DDEC.","Financial Claim","Critical","PR","EN","Yes","","",""],
    ["LIVE-BL-A60-007","Live","blog/act-60-export-services-guide","Annual Costs","Total annual compliance: $10,000-$18,000/year.","Financial Claim","High","General","EN","Yes","","",""],

    # BLOG: SENATE BILL 773
    ["LIVE-BL-SB-001","Live","blog/senate-bill-773","SB 773","Senate approved SB 773 on January 27, 2026 — 46-section overhaul of Law 219-2012.","Statutory Citation","Critical","PR","EN","Yes","","Status may have changed since March 2026.","Check current legislative status"],
    ["LIVE-BL-SB-002","Live","blog/senate-bill-773","Prior Law","Article 7 declared trusts 'shall be irrevocable.'","Statutory Citation","Critical","PR","EN","No","","",""],
    ["LIVE-BL-SB-003","Live","blog/senate-bill-773","New Default","If trust instrument silent on revocability, trust now presumed revocable. Complete reversal.","Legal Requirement","Critical","PR","EN","Yes","","Only if SB 773 enacted. Check status.","Verify enactment"],
    ["LIVE-BL-SB-004","Live","blog/senate-bill-773","Creditor Rights","Revocable trust offers no asset shield while settlor alive. Full protection only after irrevocability.","Legal Requirement","Critical","PR","EN","Yes","","",""],
    ["LIVE-BL-SB-005","Live","blog/senate-bill-773","Co-Trustee","Co-trustee liability shifts from joint-and-several to proportional.","Legal Requirement","Critical","PR","EN","Yes","","Only if SB 773 enacted.","Verify enactment"],
    ["LIVE-BL-SB-006","Live","blog/senate-bill-773","Law 241-2014","Law 241-2014 allowed Act 22 decree holders to create revocable trusts.","Statutory Citation","Critical","PR","EN","No","","",""],
    ["LIVE-BL-SB-007","Live","blog/senate-bill-773","Status","Bill endorsed by Colegio de Abogados and Asociacion de Bancos de PR.","Professional Claim","High","PR","EN","Yes","","",""],
    ["LIVE-BL-SB-008","Live","blog/senate-bill-773","Status","Legislation moving to House of Representatives.","Procedural Claim","High","PR","EN","Yes","","May have moved further by now.","Check current status"],

    # BLOG: COMPANY FORMATION
    ["LIVE-BL-CF-001","Live","blog/company-formation-pr","LLC","PR LLC provides limited liability, flexible management, pass-through taxation by default.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BL-CF-002","Live","blog/company-formation-pr","C Corp Act 60","4% rate at corporate level, distributions 100% exempt for bona fide PR residents.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-BL-CF-003","Live","blog/company-formation-pr","Partnership","General partners face unlimited personal liability.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BL-CF-004","Live","blog/company-formation-pr","Tax System","PR corporations not subject to U.S. federal income tax on PR-source income.","Tax Rate","Critical","Both","EN","No","","",""],
    ["LIVE-BL-CF-005","Live","blog/company-formation-pr","Community Property","PR is community property jurisdiction. Spouse may have automatic interest in business assets.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BL-CF-006","Live","blog/company-formation-pr","Tax Summary","C Corp with Act 60: ~4% effective. Without: 18.5%-37.5%. LLC pass-through: 0%-33%.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-BL-CF-007","Live","blog/company-formation-pr","Costs","Dept of State: $150-$350. EIN: Free. Hacienda: $50-$100. Total: $2,000-$6,000. Timeline: 2-4 weeks.","Financial Claim","High","PR","EN","Yes","","",""],
    ["LIVE-BL-CF-008","Live","blog/company-formation-pr","CRIM","Businesses with personal property must register with CRIM even without real property.","Legal Requirement","High","PR","EN","No","","",""],

    # BLOG: LIVING VS IRREVOCABLE TRUST
    ["LIVE-BL-LI-001","Live","blog/living-vs-irrevocable-trust","Living Trust","Living trust does not provide asset protection or significant tax advantages during lifetime.","Legal Requirement","High","PR","EN","No","","",""],
    ["LIVE-BL-LI-002","Live","blog/living-vs-irrevocable-trust","Irrevocable","PR trust law primarily governed by Law 219-2012.","Statutory Citation","High","PR","EN","No","","",""],
    ["LIVE-BL-LI-003","Live","blog/living-vs-irrevocable-trust","Forced Heirship","Under PR Civil Code, you cannot freely disinherit your children or descendants.","Legal Requirement","Critical","PR","EN","No","","",""],
    ["LIVE-BL-LI-004","Live","blog/living-vs-irrevocable-trust","Forced Heirship","Legitima (1/3), mejora (1/3), free disposal (1/3).","Legal Requirement","Critical","PR","EN","No","","CONTRADICTION: Should be 1/2 legitima + 1/2 libre disposicion under Ley 55-2020.","Must fix — wrong law cited"],
    ["LIVE-BL-LI-005","Live","blog/living-vs-irrevocable-trust","Forced Heirship","You cannot use a trust to circumvent the legitima. Heirs can challenge in court.","Legal Requirement","Critical","PR","EN","No","","",""],
    ["LIVE-BL-LI-006","Live","blog/living-vs-irrevocable-trust","Act 60","If domiciled in PR, legitima rules govern PR assets regardless of prior residence.","Jurisdictional Claim","Critical","PR","EN","No","","",""],
    ["LIVE-BL-LI-007","Live","blog/living-vs-irrevocable-trust","Act 60","A mainland revocable trust may not function as intended under PR's legal system.","Jurisdictional Claim","High","Both","EN","No","","",""],
    ["LIVE-BL-LI-008","Live","blog/living-vs-irrevocable-trust","Community Property","PR is community property jurisdiction, affects what can be transferred into trust.","Legal Requirement","High","PR","EN","No","","",""],

    # ACT 60 RESOURCE CENTER
    ["LIVE-RC-001","Live","riefkohllaw.com/act-60-resource-center","Residency","Act 60 benefits depend on bona fide residency test under IRC S937(a).","Statutory Citation","Critical","Federal","EN","No","","",""],
    ["LIVE-RC-002","Live","riefkohllaw.com/act-60-resource-center","Income Sourcing","Act 60 benefits apply only to PR-source income under IRC S937(b).","Statutory Citation","Critical","Federal","EN","No","","",""],
    ["LIVE-RC-003","Live","riefkohllaw.com/act-60-resource-center","Working Days","Where you physically perform work determines tax benefit. Mainland travel creates U.S.-source income.","Tax Rate","Critical","Federal","EN","No","","",""],
    ["LIVE-RC-004","Live","riefkohllaw.com/act-60-resource-center","Software","Software income classification under Reg. S1.861-18 determines Act 60 eligibility.","Statutory Citation","Critical","Federal","EN","No","","",""],
    ["LIVE-RC-005","Live","riefkohllaw.com/act-60-resource-center","Chapter 3","Chapter 3 offers 4% corporate tax rate on eligible export services income.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["LIVE-RC-006","Live","riefkohllaw.com/act-60-resource-center","Safe Harbor","Year-of-move safe harbor under Treas. Reg. S1.937-1(f), three conditions.","Statutory Citation","Critical","Federal","EN","No","","",""],
    ["LIVE-RC-007","Live","riefkohllaw.com/act-60-resource-center","NRNC","PR residents born in U.S. territory classified as NRNC for estate and gift tax.","Tax Rate","Critical","Federal","EN","No","","",""],

    # SPANISH PAGES
    ["LIVE-ES-001","Live","riefkohllaw.com/espanol","Hero","Experience at DLA Piper, LUMA Energy, and federal courts.","Professional Claim","High","General","ES","No","","",""],
    ["LIVE-ES-002","Live","riefkohllaw.com/espanol","Forced Heirship","El Codigo Civil impone la legitima — una porcion forzosa.","Legal Requirement","Critical","PR","ES","No","","",""],
    ["LIVE-ES-003","Live","riefkohllaw.com/espanol","Act 60","Hacienda y Oficina de Incentivos exigen cumplimiento continuo. Error = revocacion del decreto.","Legal Requirement","Critical","PR","ES","No","","",""],
    ["LIVE-ES-004","Live","riefkohllaw.com/espanol","Bio","Oficial juridico en Tribunal Supremo de PR y Tribunal Federal, transacciones de mas de $100M en LUMA.","Professional Claim","Critical","General","ES","No","","",""],
    ["LIVE-ES-005","Live","riefkohllaw.com/espanol-servicios","Pricing","Nivel 1: $7,500-$12,000. Nivel 2: $12,000-$18,000. Nivel 3: $18,000-$30,000. Nivel 4: $30,000-$50,000+.","Financial Claim","High","General","ES","Yes","","",""],
    ["LIVE-ES-006","Live","riefkohllaw.com/espanol-servicios","Trust Law","Fideicomisos irrevocables bajo la Ley 219.","Statutory Citation","High","PR","ES","No","","",""],
    ["LIVE-ES-007","Live","riefkohllaw.com/espanol-servicios","Bankruptcy","Quiebra bajo el Titulo 11 del Codigo federal.","Statutory Citation","High","Federal","ES","No","","",""],
    ["LIVE-ES-008","Live","riefkohllaw.com/espanol-servicios","Business","LLC: $1,500. Corp: $2,000. Partnership: $2,500. Governance: $3,500.","Financial Claim","High","General","ES","Yes","","",""],
    ["LIVE-ES-009","Live","riefkohllaw.com/espanol-servicios","Transactional","Contratos: $750. Gubernamentales: $3,000. Ley 60: $5,000.","Financial Claim","High","General","ES","Yes","","",""],
    ["LIVE-HP-001","Live","riefkohllaw.com/home-project","Hero","Backed by experience at DLA Piper, LUMA Energy, and the federal courts.","Professional Claim","High","General","EN","No","","",""],
    ["LIVE-HP-002","Live","riefkohllaw.com/home-project","Differentiator","From $100M+ transactions to small business formations.","Professional Claim","High","General","EN","No","","",""],
]

# --- LOCAL FILES: Resources, Homepage, SEO, Email, Lead Magnets, Directory (147 from agent) ---
local = [
    # RESOURCES-FIX.HTML
    ["RES-001","Local","resources-fix.html","Estate Planning FAQ","Puerto Rico trusts are governed by the Puerto Rico Trust Act (Ley 219-2012).","Statutory Citation","Medium","PR","EN","No","","",""],
    ["RES-002","Local","resources-fix.html","Forced Heirship FAQ","Under the 2020 Civil Code (Ley 55-2020), the legitimate portion (legitima) is one-half of the estate, reserved for descendants.","Statutory Citation","Critical","PR","EN","No","","This is the CORRECT current law. Contradicts live estate planning page which says 2/3.",""],
    ["RES-003","Local","resources-fix.html","Forced Heirship FAQ","The remaining half is freely disposable.","Legal Requirement","High","PR","EN","No","","",""],
    ["RES-004","Local","resources-fix.html","Probate FAQ","Puerto Rico's Trust Act permits both revocable and irrevocable trusts.","Statutory Citation","High","PR","EN","No","","CONTRADICTION with EMAIL-001/002 which say all PR trusts are irrevocable.","Resolve inconsistency"],
    ["RES-005","Local","resources-fix.html","Probate FAQ","Once established, the grantor generally cannot modify or revoke without beneficiaries' consent.","Trust Requirement","High","PR","EN","No","","",""],
    ["RES-006","Local","resources-fix.html","Wills Subpage","Wills must comply with Civil Code formalities, including being executed before a notary public and witnesses.","Procedural Claim","Critical","PR","EN","No","","CONTRADICTS CONV-030: witnesses no longer required under Art. 1644.","Must fix"],
    ["RES-007","Local","resources-fix.html","Wills Subpage","The 2020 Civil Code (Ley 55-2020) eliminated closed wills (testamento cerrado).","Statutory Citation","High","PR","EN","No","","",""],
    ["RES-008","Local","resources-fix.html","Wills Subpage","Legacy closed wills executed before November 28, 2020 may still be valid under transitional rules.","Legal Requirement","High","PR","EN","No","","",""],
    ["RES-009","Local","resources-fix.html","Forced Heirship Subpage","Legitima (one-half): Reserved for forced heirs (descendants). Must be distributed equally among all children or descendants.","Legal Requirement","Critical","PR","EN","No","","",""],
    ["RES-010","Local","resources-fix.html","Forced Heirship Subpage","The forced portion was reduced from two-thirds to one-half by Ley 55-2020.","Statutory Citation","High","PR","EN","No","","",""],
    ["RES-011","Local","resources-fix.html","Forced Heirship Subpage","The 2020 Code elevated surviving spouse from usufruct-only rights to full forced heir status, sharing legitima equally with children.","Statutory Citation","Critical","PR","EN","No","","",""],
    ["RES-012","Local","resources-fix.html","Forced Heirship Subpage","Life insurance proceeds payable to named beneficiaries generally fall outside the estate for forced heirship.","Legal Requirement","Critical","PR","EN","No","","",""],
    ["RES-013","Local","resources-fix.html","5 Mistakes Subpage","U.S. citizens in Puerto Rico are subject to federal estate tax on worldwide assets.","Legal Requirement","Critical","Both","EN","No","","",""],
    ["RES-014","Local","resources-fix.html","5 Mistakes Subpage","Act 60 income tax exemptions do not extend to estate and gift taxes.","Legal Requirement","Critical","Both","EN","No","","",""],
    ["RES-015","Local","resources-fix.html","5 Mistakes Subpage","A mainland trust does not override Puerto Rico's forced heirship rules.","Legal Requirement","Critical","PR","EN","No","","",""],
    ["RES-016","Local","resources-fix.html","Trust vs Will","Dynasty trusts can last up to 70 years from the date the trust takes effect.","Trust Requirement","High","PR","EN","No","","",""],
    ["RES-017","Local","resources-fix.html","Trust vs Will","Cost: Will $1,500-$3,000. Revocable trust $7,500-$15,000. Irrevocable trust $15,000-$50,000+.","Financial Claim","Critical","PR","EN","Yes","","",""],
    ["RES-018","Local","resources-fix.html","Asset Protection","Law 219-2012 provides modern trust statute with strong creditor protections and long duration (up to 70 years).","Statutory Citation","High","PR","EN","No","","",""],
    ["RES-019","Local","resources-fix.html","Complete Guide","Federal estate tax exemption scheduled to sunset to ~$7M in 2026.","Tax Rate","Critical","Federal","EN","Yes","","WRONG: OBBBA made $15M permanent per CONV-031.","Must fix — outdated"],
    ["RES-020","Local","resources-fix.html","Declaratoria (ES)","Caso simple: 6-12 meses. Moderado: 12-18 meses. Complejo: 18 meses a varios anos.","Procedural Claim","Medium","PR","ES","No","","",""],
    ["RES-021","Local","resources-fix.html","Declaratoria (ES)","30 a 60 dias para que acreedores presenten reclamaciones.","Procedural Claim","High","PR","ES","No","","",""],
    ["RES-022","Local","resources-fix.html","Ley 60 Guia (ES)","4% sobre servicios de exportacion (Capitulo 3).","Tax Rate","Critical","PR","ES","Yes","","",""],
    ["RES-023","Local","resources-fix.html","Ley 60 Guia (ES)","Exencion federal sobre ingreso de fuente PR (IRC S933).","Statutory Citation","Critical","Both","ES","No","","",""],
    ["RES-024","Local","resources-fix.html","Ley 60 Guia (ES)","Los beneficios de Ley 60 no se extienden al impuesto sobre caudal relicto.","Legal Requirement","Critical","Both","ES","No","","",""],
    ["RES-025","Local","resources-fix.html","Ley 60 Guia (ES)","Presencia fisica: Al menos 183 dias en Puerto Rico.","Legal Requirement","Critical","Both","ES","Yes","","",""],

    # EMAIL WELCOME SEQUENCE
    ["EMAIL-001","Local","email-welcome-sequence.md","Email 1","All Puerto Rico trusts are irrevocable.","Legal Requirement","Critical","PR","EN","No","","CONTRADICTION: RES-004 says PR Trust Act permits both revocable and irrevocable. SB 773 adds revocable default.","Must fix — inaccurate"],
    ["EMAIL-002","Local","email-welcome-sequence.md","Email 1","There is no revocable living trust under Puerto Rico law.","Legal Requirement","Critical","PR","EN","No","","WRONG: Even before SB 773, Law 241-2014 allowed revocable trusts for Act 22 decree holders.","Must fix — inaccurate"],
    ["EMAIL-003","Local","email-welcome-sequence.md","Email 1","PR irrevocable trusts can be structured with significant flexibility — power to amend terms, change trustees, add/remove beneficiaries.","Trust Requirement","Critical","PR","EN","No","","",""],
    ["EMAIL-004","Local","email-welcome-sequence.md","Email 2","PR trust law (Ley 219-2012) treats trust as autonomous estate (patrimonio autonomo), a separate legal person.","Statutory Citation","High","PR","EN","No","","",""],
    ["EMAIL-005","Local","email-welcome-sequence.md","Email 2","PR trust property belongs to the trust entity. Creditors generally cannot reach it.","Legal Requirement","Critical","PR","EN","No","","",""],
    ["EMAIL-006","Local","email-welcome-sequence.md","Email 2","PR has forced heirship (legitima). No trust or will can override it.","Legal Requirement","Critical","PR","EN","No","","",""],
    ["EMAIL-007","Local","email-welcome-sequence.md","Email 3","Fideicomiso El Puente GNR v. Asociacion de Residentes (2024): trust can sue in its own name.","Case Citation","High","PR","EN","No","","Verify case name, year, holding.",""],
    ["EMAIL-008","Local","email-welcome-sequence.md","Email 3","Gil Ensenat v. Marini Roman (2009): divorcing spouse's interest in trust funded with community property.","Case Citation","High","PR","EN","No","","Verify case name, year, holding.",""],
    ["EMAIL-009","Local","email-welcome-sequence.md","Email 4","All inter vivos trusts must be executed before a notary and registered within 30 days.","Legal Requirement","Critical","PR","EN","No","","Verify 30-day deadline.",""],
    ["EMAIL-010","Local","email-welcome-sequence.md","Email 4","Registered with ODIN (Special Trust Registry).","Legal Requirement","High","PR","EN","No","","",""],

    # LEAD MAGNETS
    ["LEAD-001","Local","lead-magnet-act60-trusts.md","Section 1","Act 60 decree exempts from income taxes on PR-sourced capital gains, interest, dividends.","Legal Requirement","Critical","Both","EN","No","","",""],
    ["LEAD-002","Local","lead-magnet-act60-trusts.md","Section 1","Does NOT exempt estate from federal estate tax or PR estate tax.","Legal Requirement","Critical","Both","EN","No","","",""],
    ["LEAD-003","Local","lead-magnet-act60-trusts.md","Section 1","Federal exemption $13.61M in 2026, scheduled to drop to ~$7M.","Tax Rate","Critical","Federal","EN","Yes","","WRONG: OBBBA made $15M permanent. Must update.","Must fix — wrong number"],
    ["LEAD-004","Local","lead-magnet-act60-trusts.md","Section 1","Heirs could face 40% estate tax rate.","Tax Rate","Critical","Federal","EN","Yes","","",""],
    ["LEAD-005","Local","lead-magnet-act60-trusts.md","Section 2","PR law reserves one-half for forced heirs — children and surviving spouse.","Legal Requirement","Critical","PR","EN","No","","Correct under Ley 55-2020.",""],
    ["LEAD-006","Local","lead-magnet-act60-trusts.md","Section 3","Before 2012, PR lacked comprehensive trust statute. Law 219-2012 created framework for both revocable and irrevocable trusts.","Statutory Citation","Critical","PR","EN","No","","",""],
    ["LEAD-007","Local","lead-magnet-act60-trusts.md","Section 5","Transferring assets to irrevocable trust after lawsuit/creditor claim can be fraudulent transfer.","Legal Requirement","High","PR","EN","No","","",""],
    ["LEAD-008","Local","lead-magnet-act60-trusts.md","Section 5","Federal estate tax exemption historically high but scheduled to decrease.","Tax Rate","Critical","Federal","EN","Yes","","WRONG: OBBBA made it permanent.","Must fix"],
    ["LEAD-009","Local","lead-magnet-estate-planning-checklist.md","Section 5","Federal exemption $13.61M in 2026.","Tax Rate","Critical","Federal","EN","Yes","","WRONG: Should be $15M per OBBBA.","Must fix"],

    # LEGAL-CONTENT-FIXES.HTML (the corrective patches themselves)
    ["CONV-001","Local","legal-content-fixes.html","ULLCA Fix","PR LLCs governed by Chapter XIX of General Corporations Act of 2009, NOT the ULLCA.","Statutory Citation","Critical","PR","EN","No","","This CORRECTS the live business formation page.",""],
    ["CONV-002","Local","legal-content-fixes.html","HB 505 Sunset","0% capital gains rate only for applications before Dec 31, 2026. After: 4% rate.","Tax Rate","Critical","PR","EN","Yes","","",""],
    ["CONV-003","Local","legal-content-fixes.html","HB 505","Existing decree holders grandfathered under current terms. Program extended through 2055.","Statutory Citation","Critical","PR","EN","Yes","","",""],
    ["CONV-004","Local","legal-content-fixes.html","Trust Registry","Every PR trust must be registered in Special Registry of Trusts. Failure = absolute nullity.","Legal Requirement","Critical","PR","EN","No","","",""],
    ["CONV-005","Local","legal-content-fixes.html","Forced Heirship","Ley 55-2020 reduced legitima from 2/3 to 1/2, eliminated mejora, added spouse as forced heir.","Statutory Citation","Critical","PR","EN","No","","This is the correction applied to live pages.",""],
    ["CONV-006","Local","legal-content-fixes.html","Wills Fix","Witnesses no longer required for open wills under Art. 1644 unless requested.","Statutory Citation","Critical","PR","EN","No","","",""],
    ["CONV-007","Local","legal-content-fixes.html","OBBBA","Estate tax exemption $15M per individual, made permanent with inflation indexing.","Tax Rate","Critical","Federal","EN","Yes","","This is the correction — but lead magnets/resources still show wrong number.",""],
    ["CONV-008","Local","legal-content-fixes.html","NRNC","NRNCs subject to unified credit of only $13,000 (exemption ~$60,000).","Tax Rate","Critical","Federal","EN","No","","",""],
    ["CONV-009","Local","legal-content-fixes.html","Charitable Donation","$10,000 annual — at least $5,000 to CECFL-approved orgs for eradicating child poverty.","Legal Requirement","Critical","PR","EN","Yes","","",""],
    ["CONV-010","Local","legal-content-fixes.html","6-Year Rule","Post-2026 applicants must show not a PR resident for 6 years prior.","Legal Requirement","Critical","PR","EN","Yes","","",""],
    ["CONV-011","Local","legal-content-fixes.html","Gift Tax","$19,000 annual exclusion for 2026.","Tax Rate","High","Federal","EN","Yes","","",""],
    ["CONV-012","Local","legal-content-fixes.html","Gelpi Fix","Judge Gelpi elevated to First Circuit in October 2021.","Professional Claim","Medium","Federal","EN","No","","",""],

    # DIRECTORY LISTINGS
    ["DIR-001","Local","directory-listings-content.md","Short Bio","Over 8 years of legal experience.","Professional Claim","High","General","EN","Yes","","Time-sensitive. Must update periodically.",""],
    ["DIR-002","Local","directory-listings-content.md","Short Bio","$100M+ in transactions across corporate law, government procurement, trust administration.","Financial Claim","Critical","General","EN","No","","Unverifiable aggregate claim.",""],
    ["DIR-003","Local","directory-listings-content.md","Short Bio","Previously practiced at DLA Piper.","Professional Claim","High","General","EN","No","","",""],
    ["DIR-004","Local","directory-listings-content.md","Short Bio","Corporate counsel at LUMA Energy.","Professional Claim","High","General","EN","No","","",""],
    ["DIR-005","Local","directory-listings-content.md","Short Bio","J.D. summa cum laude from UPR School of Law.","Professional Claim","Critical","General","EN","No","","",""],
    ["DIR-006","Local","directory-listings-content.md","Short Bio","Admitted to practice in PR and before U.S. District Court for D.P.R.","Professional Claim","Critical","General","EN","Yes","","Check if D.C. and First Circuit should also be listed.",""],
    ["DIR-007","Local","directory-listings-content.md","Long Bio","Completed judicial clerkships in federal courts.","Professional Claim","High","General","EN","No","","",""],

    # SEO META DESCRIPTIONS
    ["SEO-001","Local","seo-fixes.html","About meta","Attorney Hans Riefkohl brings 8+ years of experience from DLA Piper and LUMA Energy.","Professional Claim","High","General","EN","Yes","","Time-sensitive claim.","Update periodically"],
    ["SEO-002","Local","seo-fixes.html","Article 6 meta","Puerto Rico's autonomous estate doctrine provides asset protection no mainland DAPT can match.","Comparative Claim","Critical","Both","EN","No","","Bold comparative claim — verify legal accuracy.",""],
    ["SEO-003","Local","seo-fixes.html","Article 11 meta","Trust costs: Attorney fees $1,500-$35,000+, notarial fees, ODIN registration.","Financial Claim","Critical","PR","EN","Yes","","",""],
    ["SEO-004","Local","seo-fixes.html","Article 7 meta","Special needs trusts: Medicaid's any-circumstances test, PR's Plan Vital changes everything.","Statutory Citation","High","Both","EN","Yes","","Verify Plan Vital still current.",""],
    ["SEO-005","Local","seo-fixes.html","Article 9 meta","HB 505 and the OBBBA's impact on PR trust planning for Act 60 decree holders.","Statutory Citation","High","Both","EN","Yes","","",""],
    ["SEO-006","Local","seo-fixes.html","Article 12 meta","The 16 statutory trustee duties under Puerto Rico's Trust Act.","Statutory Citation","High","PR","EN","No","","Verify count is exactly 16.",""],

    # CONVERSION/SIGNUP COPY
    ["CONV-CTA-001","Local","email-signup-form.html","Footer","Does Your Business Qualify for the 4% Tax Rate?","Tax Rate","Critical","PR","EN","Yes","","Implies any business could qualify. Only Chapter 3 export services.","Clarify scope"],
    ["CONV-CTA-002","Local","email-signup-form.html","Footer (ES)","Su negocio cualifica para la tasa contributiva del 4%?","Tax Rate","Critical","PR","ES","Yes","","Same issue in Spanish.","Clarify scope"],
    ["CONV-CTA-003","Local","conversion-booster.html","Schema","Opening hours: Monday-Friday 09:00-17:00.","Professional Claim","Low","PR","EN","Yes","","Must be kept current.",""],
    ["CONV-CTA-004","Local","conversion-booster.html","Schema","273 Ponce de Leon Ave., San Juan, PR 00917.","Professional Claim","Low","PR","EN","Yes","","Must match actual address.",""],
]

all_statements = live + local

# ============================================================
# BUILD SHEETS
# ============================================================

# Sheet 1: All Statements
ws1 = setup_sheet(wb.active, "All Statements")
for i, s in enumerate(all_statements, 2):
    add_row(ws1, i, s)

# Sheet 2: CONTRADICTIONS & CRITICAL ISSUES
ws2 = setup_sheet(wb.create_sheet(), "Contradictions & Issues")
row = 2
for s in all_statements:
    if "CONTRADICTION" in s[11] or "WRONG" in s[11] or "ISSUE" in s[11] or "Must fix" in s[12]:
        add_row(ws2, row, s)
        # Highlight the notes column red
        ws2.cell(row=row, column=12).fill = contradiction_fill
        ws2.cell(row=row, column=12).font = Font(color="FFFFFF", bold=True)
        row += 1

# Sheet 3: Time-Sensitive
ws3 = setup_sheet(wb.create_sheet(), "Time-Sensitive Claims")
row = 2
for s in all_statements:
    if s[9] == "Yes":
        add_row(ws3, row, s)
        row += 1

# Sheet 4: Critical Severity
ws4 = setup_sheet(wb.create_sheet(), "Critical Severity")
row = 2
for s in all_statements:
    if s[6] == "Critical":
        add_row(ws4, row, s)
        row += 1

# Sheet 5: Professional Claims
ws5 = setup_sheet(wb.create_sheet(), "Professional Claims")
row = 2
for s in all_statements:
    if s[5] == "Professional Claim":
        add_row(ws5, row, s)
        row += 1

# Sheet 6: Financial & Tax Claims
ws6 = setup_sheet(wb.create_sheet(), "Financial & Tax Claims")
row = 2
for s in all_statements:
    if s[5] in ("Financial Claim", "Tax Rate"):
        add_row(ws6, row, s)
        row += 1

# Sheet 7: Summary Dashboard
ws7 = wb.create_sheet("Summary Dashboard")
ws7.column_dimensions['A'].width = 35
ws7.column_dimensions['B'].width = 15
ws7.column_dimensions['C'].width = 15

ws7.cell(row=1, column=1, value="Riefkohl Law — Liability Statement Database").font = Font(bold=True, size=14)
ws7.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True)
ws7.cell(row=3, column=1, value=f"Total Statements: {len(all_statements)}").font = Font(bold=True, size=12)
ws7.cell(row=4, column=1, value=f"  Live Site: {len(live)}  |  Local Files: {len(local)}")

# Contradictions count
contradictions = sum(1 for s in all_statements if "CONTRADICTION" in s[11] or "WRONG" in s[11] or "Must fix" in s[12])
ws7.cell(row=5, column=1, value=f"CONTRADICTIONS / MUST-FIX ISSUES: {contradictions}").font = Font(bold=True, color="FF0000", size=12)

# By Severity
ws7.cell(row=7, column=1, value="BY SEVERITY").font = Font(bold=True, size=11)
for i, h in enumerate(["Severity", "Count", "%"], 1):
    c = ws7.cell(row=8, column=i, value=h); c.font = header_font; c.fill = header_fill
for i, sev in enumerate(["Critical", "High", "Medium", "Low"], 9):
    count = sum(1 for s in all_statements if s[6] == sev)
    ws7.cell(row=i, column=1, value=sev)
    ws7.cell(row=i, column=2, value=count)
    pct = f"{count/len(all_statements)*100:.1f}%" if all_statements else "0%"
    ws7.cell(row=i, column=3, value=pct)
    fills = {"Critical": critical_fill, "High": high_fill, "Medium": medium_fill, "Low": low_fill}
    if sev in fills: ws7.cell(row=i, column=1).fill = fills[sev]

# By Category
ws7.cell(row=15, column=1, value="BY CATEGORY").font = Font(bold=True, size=11)
for i, h in enumerate(["Category", "Count", "%"], 1):
    c = ws7.cell(row=16, column=i, value=h); c.font = header_font; c.fill = header_fill
cats = {}
for s in all_statements: cats[s[5]] = cats.get(s[5], 0) + 1
for i, (cat, count) in enumerate(sorted(cats.items(), key=lambda x: -x[1]), 17):
    ws7.cell(row=i, column=1, value=cat)
    ws7.cell(row=i, column=2, value=count)
    ws7.cell(row=i, column=3, value=f"{count/len(all_statements)*100:.1f}%")

# Key metrics
ts = sum(1 for s in all_statements if s[9] == "Yes")
r = 17 + len(cats) + 1
ws7.cell(row=r, column=1, value=f"Time-Sensitive Claims: {ts} ({ts/len(all_statements)*100:.1f}%)").font = Font(bold=True)
ws7.cell(row=r+1, column=1, value=f"Statements needing immediate action: {contradictions}").font = Font(bold=True, color="FF0000")

# Save
out = "/Users/hansriefkohl/Downloads/Claude Code/Riefkohl Law updates/liability-statement-database.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total: {len(all_statements)} | Live: {len(live)} | Local: {len(local)}")
print(f"Critical: {sum(1 for s in all_statements if s[6]=='Critical')}")
print(f"Contradictions/Must-Fix: {contradictions}")
print(f"Time-Sensitive: {ts}")
