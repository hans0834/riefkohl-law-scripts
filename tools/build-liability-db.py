#!/usr/bin/env python3
"""
Riefkohl Law — Liability Statement Database Builder
Extracts every factual/legal statement from the website that could create liability if inaccurate.
Outputs to liability-statement-database.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# STYLING
# ============================================================
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
high_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
medium_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
low_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
wrap_alignment = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

HEADERS = [
    "ID", "Source", "Source URL/File", "Page/Section", "Statement (Verbatim)",
    "Category", "Severity", "Jurisdiction", "Language", "Time-Sensitive?",
    "Verified", "Notes", "Action Needed"
]

COL_WIDTHS = [14, 8, 45, 30, 80, 20, 10, 12, 5, 14, 10, 30, 15]

def setup_sheet(ws, title):
    ws.title = title
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    return ws

def add_row(ws, row_num, data):
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        # Color severity column
        if col_idx == 7:
            if val == "Critical":
                cell.fill = critical_fill
            elif val == "High":
                cell.fill = high_fill
            elif val == "Medium":
                cell.fill = medium_fill
            elif val == "Low":
                cell.fill = low_fill

# ============================================================
# SHEET 1: LIVE SITE STATEMENTS
# ============================================================
ws1 = setup_sheet(wb.active, "Live Site Statements")

# All statements extracted from live Squarespace pages
live_statements = [
    # ---- ESTATE PLANNING PAGE ----
    ["LIVE-EP-001", "Live", "riefkohllaw.com/estate-planning", "Estate Planning - Intro",
     "Puerto Rico operates under a civil law system inherited from Spanish legal traditions, unlike the common law system used in most U.S. states.",
     "Jurisdictional Claim", "Medium", "PR", "EN", "No", "", "", ""],
    ["LIVE-EP-002", "Live", "riefkohllaw.com/estate-planning", "Forced Heirship",
     "Under Articles 739 through 762 of the Puerto Rico Civil Code, two-thirds of your estate is reserved for your forced heirs, typically your children or descendants.",
     "Statutory Citation", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-EP-003", "Live", "riefkohllaw.com/estate-planning", "Forced Heirship",
     "Your estate is divided into three portions: the legitima estricta (one-third reserved equally for forced heirs), the mejora (one-third that can be distributed among forced heirs as you choose), and the libre disposicion (one-third that you can freely distribute to anyone).",
     "Legal Requirement", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-EP-004", "Live", "riefkohllaw.com/estate-planning", "Forced Heirship",
     "This means you cannot simply disinherit your children or leave your entire estate to a single beneficiary, as is possible in many mainland states.",
     "Comparative Claim", "High", "Both", "EN", "No", "", "", ""],
    ["LIVE-EP-005", "Live", "riefkohllaw.com/estate-planning", "Wills Under PR Law",
     "Puerto Rico recognizes several types of wills, each with specific formal requirements. The most common is the open notarial will (testamento abierto), which must be executed before a notary public and witnesses.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-EP-006", "Live", "riefkohllaw.com/estate-planning", "Wills Under PR Law",
     "Puerto Rico also recognizes closed wills and holographic wills, though holographic wills have strict requirements to be valid.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-EP-007", "Live", "riefkohllaw.com/estate-planning", "Trusts",
     "Puerto Rico has its own Trust Act that governs the creation and administration of trusts on the island.",
     "Statutory Citation", "Medium", "PR", "EN", "No", "", "", ""],
    ["LIVE-EP-008", "Live", "riefkohllaw.com/estate-planning", "Probate",
     "When someone passes away in Puerto Rico, their estate typically goes through a probate process known as declaratoria de herederos.",
     "Procedural Claim", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-EP-009", "Live", "riefkohllaw.com/estate-planning", "Probate",
     "The probate process in Puerto Rico can take anywhere from six months to several years.",
     "Procedural Claim", "Medium", "PR", "EN", "No", "", "", ""],
    ["LIVE-EP-010", "Live", "riefkohllaw.com/estate-planning", "Asset Protection",
     "Puerto Rico offers several asset protection tools including irrevocable trusts, private interest foundations (fundaciones de interes privado), proper business entity structuring through LLCs and corporations, and insurance planning.",
     "Legal Requirement", "Medium", "PR", "EN", "No", "", "", ""],

    # ---- TRUSTS PAGE ----
    ["LIVE-TR-001", "Live", "riefkohllaw.com/puerto-rico-trusts", "Understanding PR Trust Law",
     "Puerto Rico trusts are governed by the Puerto Rico Trust Act (Ley de Fideicomisos), which provides a flexible and modern trust framework rooted in civil law traditions.",
     "Statutory Citation", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-TR-002", "Live", "riefkohllaw.com/puerto-rico-trusts", "Understanding PR Trust Law",
     "Unlike many U.S. mainland jurisdictions that follow common law trust principles, Puerto Rico's trust legislation blends civil law concepts with elements from the Uniform Trust Code, creating a hybrid system that offers distinct advantages.",
     "Comparative Claim", "Medium", "Both", "EN", "No", "", "", ""],
    ["LIVE-TR-003", "Live", "riefkohllaw.com/puerto-rico-trusts", "Revocable Living Trusts",
     "A revocable living trust allows you to maintain full control over your assets during your lifetime while providing seamless transfer to beneficiaries upon death.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-TR-004", "Live", "riefkohllaw.com/puerto-rico-trusts", "Irrevocable Trusts",
     "Once established, an irrevocable trust generally cannot be modified or revoked without the consent of the beneficiaries.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-TR-005", "Live", "riefkohllaw.com/puerto-rico-trusts", "Special Needs Trusts",
     "A special needs trust enables you to provide for a disabled family member without jeopardizing their eligibility for government benefits such as Medicaid or Supplemental Security Income.",
     "Legal Requirement", "Critical", "Both", "EN", "No", "", "", ""],
    ["LIVE-TR-006", "Live", "riefkohllaw.com/puerto-rico-trusts", "Dynasty Trusts",
     "Puerto Rico's favorable trust laws make it an attractive jurisdiction for establishing multi-generational wealth preservation structures.",
     "Comparative Claim", "Medium", "PR", "EN", "No", "", "", ""],
    ["LIVE-TR-007", "Live", "riefkohllaw.com/puerto-rico-trusts", "Private Interest Foundations",
     "In addition to traditional trusts, Puerto Rico law permits the creation of private interest foundations, which function similarly to trusts but operate as separate legal entities.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-TR-008", "Live", "riefkohllaw.com/puerto-rico-trusts", "FAQ - Creating a Trust",
     "A trust is created through a written trust agreement (escritura de fideicomiso) executed before a notary public.",
     "Procedural Claim", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-TR-009", "Live", "riefkohllaw.com/puerto-rico-trusts", "FAQ - Asset Protection",
     "Irrevocable trusts in Puerto Rico can provide significant asset protection from creditors, lawsuits, and other claims, provided the trust is properly structured and funded in a timely manner before any claims arise.",
     "Legal Requirement", "Critical", "PR", "EN", "No", "", "", ""],

    # ---- BUSINESS FORMATION PAGE ----
    ["LIVE-BF-001", "Live", "riefkohllaw.com/business-formation", "Choosing Entity",
     "Puerto Rico recognizes several business entity types under its General Corporations Act (Ley General de Corporaciones) and other statutes.",
     "Statutory Citation", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BF-002", "Live", "riefkohllaw.com/business-formation", "LLCs",
     "LLCs in Puerto Rico are governed by the Puerto Rico Uniform Limited Liability Company Act and can be structured as single-member or multi-member entities.",
     "Statutory Citation", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BF-003", "Live", "riefkohllaw.com/business-formation", "Corporations",
     "The island's General Corporations Act closely mirrors Delaware corporate law, making it familiar to mainland practitioners.",
     "Comparative Claim", "Medium", "Both", "EN", "No", "", "", ""],
    ["LIVE-BF-004", "Live", "riefkohllaw.com/business-formation", "Name Reservation",
     "Once confirmed, we can reserve the business name for up to 120 days while preparing the formation documents.",
     "Procedural Claim", "High", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BF-005", "Live", "riefkohllaw.com/business-formation", "SURI Registration",
     "Registration with Puerto Rico's Unified Internal Revenue System (SURI) is required for local tax compliance, including sales and use tax (IVU) collection obligations.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BF-006", "Live", "riefkohllaw.com/business-formation", "Municipal Licenses",
     "Each of Puerto Rico's 78 municipalities has its own licensing requirements.",
     "Jurisdictional Claim", "Medium", "PR", "EN", "No", "", "", ""],
    ["LIVE-BF-007", "Live", "riefkohllaw.com/business-formation", "Tax Considerations",
     "The standard corporate tax rate in Puerto Rico ranges from 18.5% to 37.5%, but significant incentives are available through various tax incentive programs.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BF-008", "Live", "riefkohllaw.com/business-formation", "Act 60 Business Incentives",
     "Eligible companies can benefit from a 4% fixed income tax rate on export service income, 100% tax exemption on distributions from earnings and profits, 75% exemption on municipal taxes, and 60% exemption on personal and real property taxes.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BF-009", "Live", "riefkohllaw.com/business-formation", "Foreign Entity",
     "If your business is already incorporated in another U.S. state or foreign jurisdiction and you want to operate in Puerto Rico, you must register as a foreign entity with the Puerto Rico Department of State.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BF-010", "Live", "riefkohllaw.com/business-formation", "FAQ - Timeline",
     "Standard processing typically takes 5-10 business days with the Department of State. Expedited processing is available. The full setup including EIN, SURI registration, and municipal licenses may take 2-6 weeks.",
     "Procedural Claim", "Medium", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BF-011", "Live", "riefkohllaw.com/business-formation", "FAQ - Registered Agent",
     "All business entities formed in Puerto Rico must maintain a registered agent with a physical address on the island to receive legal and government correspondence.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BF-012", "Live", "riefkohllaw.com/business-formation", "FAQ - Non-Resident",
     "Non-residents can form and own businesses in Puerto Rico with no residency requirement for formation. However, certain tax incentives under Act 60 may require establishing bona fide residency.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BF-013", "Live", "riefkohllaw.com/business-formation", "Partnerships",
     "In a general partnership, all partners share liability. Limited partnerships allow passive investors to limit their exposure.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BF-014", "Live", "riefkohllaw.com/business-formation", "Compliance",
     "Failure to comply can result in penalties, loss of good standing, or even involuntary dissolution of the entity.",
     "Penalty/Consequence", "High", "PR", "EN", "No", "", "", ""],

    # ---- ACT 60 PAGE ----
    ["LIVE-A60-001", "Live", "riefkohllaw.com/act-60-tax-incentives", "Understanding Act 60",
     "Act 60, formally known as the Puerto Rico Incentives Code, was enacted in 2019 to consolidate previous tax incentive laws — including Act 20 (Export Services Act) and Act 22 (Individual Investors Act) — into a single comprehensive framework.",
     "Statutory Citation", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-A60-002", "Live", "riefkohllaw.com/act-60-tax-incentives", "Understanding Act 60",
     "Act 60 benefits are granted through individual tax decrees — legally binding agreements between the applicant and the Puerto Rico government that guarantee specific tax rates and exemptions for a defined period, typically 15 years with the option to extend.",
     "Legal Requirement", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-003", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 3 Tax Benefits",
     "Qualifying export service businesses can enjoy a 4% fixed corporate income tax rate on eligible export service income, compared to the standard Puerto Rico corporate tax rate of 18.5% to 37.5%.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-004", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 3 Tax Benefits",
     "100% tax exemption on distributions from earnings and profits to shareholders.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-005", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 3 Tax Benefits",
     "75% exemption on municipal license taxes (patente municipal), reducing the effective rate to approximately 0.125% to 0.375%.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-006", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 3 Tax Benefits",
     "60% exemption on personal and real property taxes.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-007", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 3 Operations",
     "Businesses must establish a bona fide office in Puerto Rico, hire a minimum number of employees (typically at least one full-time employee in addition to the owner within the first two years).",
     "Legal Requirement", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-008", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 2 Benefits",
     "100% tax exemption on capital gains accrued after becoming a bona fide Puerto Rico resident, including gains from stocks, bonds, cryptocurrency, and other investment assets.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-009", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 2 Residency",
     "Being physically present in Puerto Rico for at least 183 days per year.",
     "Legal Requirement", "Critical", "Both", "EN", "No", "", "", ""],
    ["LIVE-A60-010", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 2 Residency",
     "Purchasing residential property in Puerto Rico within two years of the decree grant.",
     "Legal Requirement", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-011", "Live", "riefkohllaw.com/act-60-tax-incentives", "Chapter 2 Residency",
     "Making an annual charitable donation of $10,000 to qualifying Puerto Rico nonprofits.",
     "Financial Claim", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-012", "Live", "riefkohllaw.com/act-60-tax-incentives", "Application Process",
     "The typical processing time ranges from 60 to 120 days.",
     "Procedural Claim", "Medium", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-A60-013", "Live", "riefkohllaw.com/act-60-tax-incentives", "Common Mistakes",
     "Not purchasing residential property within the required two-year window.",
     "Filing Deadline", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-A60-014", "Live", "riefkohllaw.com/act-60-tax-incentives", "FAQ",
     "Benefits can be revoked for non-compliance, including failing to maintain residency, not meeting employment requirements, missing annual reports, or not making required charitable donations.",
     "Penalty/Consequence", "Critical", "PR", "EN", "No", "", "", ""],

    # ---- GOVERNMENT CONTRACTS PAGE ----
    ["LIVE-GC-001", "Live", "riefkohllaw.com/government-contracts", "Intro",
     "With billions of dollars in federal spending flowing to the island annually—including disaster recovery funds, military installations, and infrastructure projects.",
     "Financial Claim", "Medium", "Federal", "EN", "Yes", "", "", ""],
    ["LIVE-GC-002", "Live", "riefkohllaw.com/government-contracts", "Federal Procurement",
     "Federal procurement in Puerto Rico is governed by the Federal Acquisition Regulation (FAR) and agency-specific supplements.",
     "Statutory Citation", "High", "Federal", "EN", "No", "", "", ""],
    ["LIVE-GC-003", "Live", "riefkohllaw.com/government-contracts", "Federal Procurement",
     "Puerto Rico's designation as a HUBZone (Historically Underutilized Business Zone) across much of the island provides additional competitive advantages for qualifying small businesses.",
     "Jurisdictional Claim", "High", "Federal", "EN", "Yes", "", "", ""],
    ["LIVE-GC-004", "Live", "riefkohllaw.com/government-contracts", "Commonwealth Procurement",
     "Commonwealth procurement is governed by Puerto Rico's Uniform Procurement Regulation and administered through the General Services Administration (Administración de Servicios Generales, or ASG).",
     "Statutory Citation", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-GC-005", "Live", "riefkohllaw.com/government-contracts", "Registration",
     "You must obtain a Unique Entity Identifier (UEI) and register in the System for Award Management (SAM.gov).",
     "Procedural Claim", "High", "Federal", "EN", "No", "", "", ""],
    ["LIVE-GC-006", "Live", "riefkohllaw.com/government-contracts", "Bid Protests",
     "Protests must be filed within 10 days of when the basis of protest is known. GAO issues decisions within 100 days.",
     "Filing Deadline", "Critical", "Federal", "EN", "No", "", "", ""],
    ["LIVE-GC-007", "Live", "riefkohllaw.com/government-contracts", "Commonwealth Disputes",
     "The Board of Appeals for Government Procurement (Junta de Subastas) handles many Commonwealth procurement challenges.",
     "Jurisdictional Claim", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-GC-008", "Live", "riefkohllaw.com/government-contracts", "Compliance",
     "Failure to maintain compliance can result in contract termination, suspension, debarment, and potential civil or criminal liability.",
     "Penalty/Consequence", "High", "Federal", "EN", "No", "", "", ""],
    ["LIVE-GC-009", "Live", "riefkohllaw.com/government-contracts", "Davis-Bacon",
     "Davis-Bacon Act — Prevailing wage requirements for federally funded construction projects.",
     "Statutory Citation", "High", "Federal", "EN", "No", "", "", ""],
    ["LIVE-GC-010", "Live", "riefkohllaw.com/government-contracts", "Disaster Recovery",
     "Puerto Rico has received tens of billions of dollars in federal disaster recovery funding following hurricanes Maria and Fiona.",
     "Financial Claim", "Medium", "Federal", "EN", "Yes", "", "", ""],
    ["LIVE-GC-011", "Live", "riefkohllaw.com/government-contracts", "CMMC",
     "Cybersecurity Requirements (CMMC) — Evolving requirements for contractors handling controlled unclassified information.",
     "Legal Requirement", "High", "Federal", "EN", "Yes", "", "", ""],
    ["LIVE-GC-012", "Live", "riefkohllaw.com/government-contracts", "Buy American",
     "Buy American Act — Requirements for domestic sourcing of materials and products.",
     "Statutory Citation", "High", "Federal", "EN", "No", "", "", ""],

    # ---- ABOUT PAGE ----
    ["LIVE-AB-001", "Live", "riefkohllaw.com/about", "Bio",
     "Licensed to practice in Puerto Rico, Washington, D.C., the U.S. District Court for the District of Puerto Rico, and the U.S. Court of Appeals for the First Circuit.",
     "Professional Claim", "Critical", "Both", "EN", "No", "", "", ""],
    ["LIVE-AB-002", "Live", "riefkohllaw.com/about", "Bio",
     "He draws on more than a decade of experience across accounting, corporate transactions, government contracting, and complex litigation.",
     "Professional Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-AB-003", "Live", "riefkohllaw.com/about", "Bio",
     "Hans earned his Bachelor of Business Administration in Accounting and Finance, cum laude, from Villanova University.",
     "Professional Claim", "Critical", "General", "EN", "No", "", "", ""],
    ["LIVE-AB-004", "Live", "riefkohllaw.com/about", "Bio",
     "Spent time at Deloitte in Philadelphia.",
     "Professional Claim", "Critical", "General", "EN", "No", "", "", ""],
    ["LIVE-AB-005", "Live", "riefkohllaw.com/about", "Bio",
     "He also served as a notary public—a credential that carries particular weight in Puerto Rico's civil law tradition.",
     "Professional Claim", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-AB-006", "Live", "riefkohllaw.com/about", "Bio",
     "Hans earned his Juris Doctor summa cum laude from the University of Puerto Rico School of Law.",
     "Professional Claim", "Critical", "General", "EN", "No", "", "", ""],
    ["LIVE-AB-007", "Live", "riefkohllaw.com/about", "Bio",
     "He then clerked for Chief Justice Maite D. Oronoz Rodríguez at the Supreme Court of Puerto Rico.",
     "Professional Claim", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-AB-008", "Live", "riefkohllaw.com/about", "Bio",
     "Clerked for Judge Gustavo A. Gelpí at the U.S. District Court for the District of Puerto Rico.",
     "Professional Claim", "Critical", "Federal", "EN", "No", "", "", ""],
    ["LIVE-AB-009", "Live", "riefkohllaw.com/about", "Bio",
     "As a Corporate Associate at DLA Piper, he advised on public-private partnerships, cross-border transactions, and government contracting across multiple jurisdictions.",
     "Professional Claim", "Critical", "General", "EN", "No", "", "", ""],
    ["LIVE-AB-010", "Live", "riefkohllaw.com/about", "Bio",
     "At Marini Pietrantoni Muñiz—one of Puerto Rico's premier law firms—he handled complex commercial litigation.",
     "Professional Claim", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-AB-011", "Live", "riefkohllaw.com/about", "Bio",
     "As Lead Transactions Counsel for LUMA Energy, he managed the legal framework behind Puerto Rico's power grid, negotiating contracts worth hundreds of millions of dollars.",
     "Professional Claim", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-AB-012", "Live", "riefkohllaw.com/about", "Bio",
     "In 2022, Hans founded Riefkohl LLC.",
     "Professional Claim", "High", "General", "EN", "No", "", "", ""],

    # ---- SERVICES/PRICING PAGE ----
    ["LIVE-SV-001", "Live", "riefkohllaw.com/services", "Trust Pricing",
     "Tier 1: Simple Single-Beneficiary Trust — From $1,875 / $7,500 - $12,000.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-002", "Live", "riefkohllaw.com/services", "Trust Pricing",
     "Tier 2: Family Protection Trust — From $3,000 / $12,000 - $18,000.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-003", "Live", "riefkohllaw.com/services", "Trust Pricing",
     "Tier 3: Comprehensive Estate Planning Trust — From $4,500 / $18,000 - $30,000.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-004", "Live", "riefkohllaw.com/services", "Trust Pricing",
     "Tier 4: Premium/Institutional Grade — From $7,500 / $30,000 - $50,000+.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-005", "Live", "riefkohllaw.com/services", "Business Formation",
     "LLC Formation Starting at $1,500.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-006", "Live", "riefkohllaw.com/services", "Business Formation",
     "Corporation Formation Starting at $2,000.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-007", "Live", "riefkohllaw.com/services", "Transactional",
     "Contract Drafting & Review Starting at $750.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-008", "Live", "riefkohllaw.com/services", "Transactional",
     "Government Contract Support Starting at $3,000.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-009", "Live", "riefkohllaw.com/services", "Transactional",
     "Puerto Rico Act 60 Tax Incentive Services Starting at $5,000.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-010", "Live", "riefkohllaw.com/services", "Bankruptcy",
     "Creditor Representation in Bankruptcy Starting at $5,000.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-011", "Live", "riefkohllaw.com/services", "Disclaimer",
     "Admitted to practice in Puerto Rico and before the U.S. District Court for the District of Puerto Rico.",
     "Professional Claim", "Critical", "Both", "EN", "No", "", "", ""],
    ["LIVE-SV-012", "Live", "riefkohllaw.com/services", "Trust Tier 1 Details",
     "Tier 1 includes initial consultation (2-3 hours), deed preparation and review, notarization coordination, Property Registry filing coordination.",
     "Professional Claim", "Medium", "General", "EN", "Yes", "", "", ""],
    ["LIVE-SV-013", "Live", "riefkohllaw.com/services", "Trust Tier 3 Details",
     "Tax compliance roadmap (3520/3520-A guidance).",
     "Tax Rate", "High", "Federal", "EN", "No", "", "", ""],

    # ---- BLOG: COMPANY FORMATION ----
    ["LIVE-BL-CF-001", "Live", "riefkohllaw.com/blog/how-to-form-a-company...", "Entity Types - LLC",
     "A Puerto Rico LLC provides limited liability protection for its members, flexible management structures, and pass-through taxation by default.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-CF-002", "Live", "riefkohllaw.com/blog/how-to-form-a-company...", "Entity Types - C Corp",
     "For Act 60 Chapter 3 decree holders, a C corporation is often the preferred structure because the 4% fixed tax rate applies at the corporate level, and distributions to shareholders who are bona fide Puerto Rico residents are 100% tax exempt.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-CF-003", "Live", "riefkohllaw.com/blog/how-to-form-a-company...", "Entity Types - Partnership",
     "General partners face unlimited personal liability.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-CF-004", "Live", "riefkohllaw.com/blog/how-to-form-a-company...", "PR Tax System",
     "Puerto Rico has its own tax code. PR corporations are not subject to U.S. federal income tax on Puerto Rico-source income.",
     "Tax Rate", "Critical", "Both", "EN", "No", "", "", ""],
    ["LIVE-BL-CF-005", "Live", "riefkohllaw.com/blog/how-to-form-a-company...", "Community Property",
     "Puerto Rico is a community property jurisdiction. If you are married, your spouse may have an automatic interest in business assets.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-CF-006", "Live", "riefkohllaw.com/blog/how-to-form-a-company...", "Tax Summary",
     "C Corp with Act 60: 4% corporate tax, 0% on distributions = ~4% effective rate. C Corp without Act 60: 18.5%-37.5% corporate tax + tax on distributions. LLC (pass-through): Income taxed at individual rates (0%-33%).",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-CF-007", "Live", "riefkohllaw.com/blog/how-to-form-a-company...", "Costs",
     "Department of State filing: $150-$350. EIN application: Free. Hacienda registration: $50-$100. CRIM registration: $25-$50. Legal fees: $1,500-$5,000. Total: $2,000-$6,000. Timeline: 2-4 weeks.",
     "Financial Claim", "High", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-CF-008", "Live", "riefkohllaw.com/blog/how-to-form-a-company...", "CRIM",
     "Even if you do not own real property, businesses with personal property (equipment, furniture, vehicles) in Puerto Rico must register.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],

    # ---- BLOG: ACT 60 TAX GUIDE ----
    ["LIVE-BL-A60-001", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "Chapter 3",
     "Chapter 3 replaced and updated the former Act 20 (Export Services Act).",
     "Statutory Citation", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-A60-002", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "Decree Terms",
     "The decree is a binding contract between the business and the Government of Puerto Rico, typically lasting 15 years with the option to request a 15-year renewal.",
     "Legal Requirement", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-A60-003", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "Municipal Tax",
     "75% Municipal License Tax Exemption, reducing the effective rate to approximately 0.125% to 0.375%.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-A60-004", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "Tax Comparison",
     "Puerto Rico Act 60: Combined effective rate ~4%, tax on $1M = ~$40,000. California: Combined effective rate ~29.84%, tax on $1M = ~$298,400.",
     "Tax Rate", "Critical", "Both", "EN", "Yes", "", "", ""],
    ["LIVE-BL-A60-005", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "Section 933",
     "Income sourced to Puerto Rico under a valid Act 60 decree is generally not subject to U.S. federal income tax under Section 933 of the Internal Revenue Code.",
     "Statutory Citation", "Critical", "Federal", "EN", "No", "", "", ""],
    ["LIVE-BL-A60-006", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "Application Costs",
     "DDEC Filing Fee: $750. Legal Fees: $5,000-$15,000. CPA/Tax Advisor: $2,000-$5,000. Total: $8,250-$22,250.",
     "Financial Claim", "High", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-A60-007", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "Annual Compliance",
     "Pay a $5,000 annual compliance fee.",
     "Financial Claim", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-A60-008", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "Annual Costs",
     "DDEC Annual Fee: $5,000. CPA/Tax Preparation: $3,000-$8,000. Legal Compliance Review: $2,000-$5,000. Total: $10,000-$18,000/year.",
     "Financial Claim", "High", "General", "EN", "Yes", "", "", ""],
    ["LIVE-BL-A60-009", "Live", "riefkohllaw.com/blog/...act-60-export-services...", "2025-2026 Updates",
     "Increased Scrutiny on Economic Substance. Processing Timelines: Plan for 90-120 days rather than the 60-day minimum.",
     "Procedural Claim", "Medium", "PR", "EN", "Yes", "", "", ""],

    # ---- BLOG: SENATE BILL 773 ----
    ["LIVE-BL-SB-001", "Live", "riefkohllaw.com/blog/...senate-bill-773...", "SB 773",
     "On January 27, 2026, the Puerto Rico Senate approved Senate Bill 773 (P. del S. 773), a sweeping 46-section overhaul of the Puerto Rico Trust Act (Law 219-2012) that formally recognizes revocable trusts for the first time.",
     "Statutory Citation", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-SB-002", "Live", "riefkohllaw.com/blog/...senate-bill-773...", "Prior Law",
     "Article 7 declared that trusts on the island 'shall be irrevocable,' yet other provisions in the same statute contemplated revocable structures.",
     "Statutory Citation", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-SB-003", "Live", "riefkohllaw.com/blog/...senate-bill-773...", "SB 773 Changes",
     "Default Presumption of Revocability. If a trust instrument is silent on revocability, the trust is now presumed revocable. This is a complete reversal of the prior default.",
     "Legal Requirement", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-SB-004", "Live", "riefkohllaw.com/blog/...senate-bill-773...", "SB 773 Changes",
     "A revocable trust offers no asset shield while the settlor is alive. Creditors can reach trust assets as if the trust did not exist. Full asset protection only attaches once the trust becomes irrevocable by the settlor's death, incapacity, or express conversion.",
     "Legal Requirement", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-SB-005", "Live", "riefkohllaw.com/blog/...senate-bill-773...", "SB 773 Changes",
     "Co-trustee liability shifts from joint-and-several to proportional.",
     "Legal Requirement", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-SB-006", "Live", "riefkohllaw.com/blog/...senate-bill-773...", "Prior Exception",
     "Law 241-2014 further complicated matters by expressly allowing holders of Individual Resident Investor decrees to constitute revocable trusts under Puerto Rico law.",
     "Statutory Citation", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-SB-007", "Live", "riefkohllaw.com/blog/...senate-bill-773...", "Status",
     "The bill earned endorsements from the Colegio de Abogados y Abogadas de Puerto Rico and the Asociacion de Bancos de Puerto Rico.",
     "Professional Claim", "High", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-BL-SB-008", "Live", "riefkohllaw.com/blog/...senate-bill-773...", "Status",
     "We are closely monitoring this legislation as it moves to the House of Representatives.",
     "Procedural Claim", "High", "PR", "EN", "Yes", "", "", ""],

    # ---- BLOG: LIVING TRUST VS IRREVOCABLE ----
    ["LIVE-BL-LI-001", "Live", "riefkohllaw.com/blog/...living-trust-vs-irrevocable...", "Living Trust",
     "A living trust does not provide asset protection from creditors or significant tax advantages during your lifetime because you still control the assets.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-LI-002", "Live", "riefkohllaw.com/blog/...living-trust-vs-irrevocable...", "Irrevocable Trust",
     "Puerto Rico's trust law, primarily governed by Law 219-2012 (the Puerto Rico Trust Act), provides a modern and flexible framework for creating irrevocable trusts.",
     "Statutory Citation", "High", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-LI-003", "Live", "riefkohllaw.com/blog/...living-trust-vs-irrevocable...", "Forced Heirship",
     "Under Puerto Rico's Civil Code, you cannot freely disinherit your children or descendants. The law reserves a mandatory portion of your estate for your forced heirs.",
     "Legal Requirement", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-LI-004", "Live", "riefkohllaw.com/blog/...living-trust-vs-irrevocable...", "Forced Heirship",
     "The legitima (one-third). Must be divided equally among your children or descendants. You cannot redirect it. The mejora (one-third). Must go to children or descendants. The free disposal portion (one-third). The only portion you can leave to anyone you choose.",
     "Legal Requirement", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-LI-005", "Live", "riefkohllaw.com/blog/...living-trust-vs-irrevocable...", "Forced Heirship",
     "You cannot use a trust to circumvent the legitima. If a trust violates forced heirship rules, affected heirs can challenge it in court.",
     "Legal Requirement", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-LI-006", "Live", "riefkohllaw.com/blog/...living-trust-vs-irrevocable...", "Act 60 Considerations",
     "Forced heirship applies to you. If you are domiciled in Puerto Rico, the legitima rules govern your PR assets, regardless of where you lived before.",
     "Jurisdictional Claim", "Critical", "PR", "EN", "No", "", "", ""],
    ["LIVE-BL-LI-007", "Live", "riefkohllaw.com/blog/...living-trust-vs-irrevocable...", "Act 60 Considerations",
     "A revocable trust created under California or New York law may not function as intended under Puerto Rico's legal system.",
     "Jurisdictional Claim", "High", "Both", "EN", "No", "", "", ""],
    ["LIVE-BL-LI-008", "Live", "riefkohllaw.com/blog/...living-trust-vs-irrevocable...", "Act 60 Considerations",
     "Puerto Rico is a community property jurisdiction, affecting what can be transferred into a trust.",
     "Legal Requirement", "High", "PR", "EN", "No", "", "", ""],

    # ---- ACT 60 RESOURCE CENTER ----
    ["LIVE-RC-001", "Live", "riefkohllaw.com/act-60-resource-center", "Residency",
     "Act 60 benefits depend on passing the federal bona fide residency test under IRC §937(a). Learn the presence test, tax home test, and closer connection test requirements.",
     "Statutory Citation", "Critical", "Federal", "EN", "No", "", "", ""],
    ["LIVE-RC-002", "Live", "riefkohllaw.com/act-60-resource-center", "Income Sourcing",
     "Act 60 tax benefits apply only to Puerto Rico-source income. Learn how compensation, investment income, capital gains, and business income are sourced under IRC §937(b).",
     "Statutory Citation", "Critical", "Federal", "EN", "No", "", "", ""],
    ["LIVE-RC-003", "Live", "riefkohllaw.com/act-60-resource-center", "Working Days",
     "For services income, where you physically perform work determines your tax benefit. Learn how the working-days allocation works and why mainland travel creates U.S.-source income.",
     "Tax Rate", "Critical", "Federal", "EN", "No", "", "", ""],
    ["LIVE-RC-004", "Live", "riefkohllaw.com/act-60-resource-center", "Software Income",
     "How classification of software income under Reg. §1.861-18 determines whether it qualifies for Act 60 benefits.",
     "Statutory Citation", "Critical", "Federal", "EN", "No", "", "", ""],
    ["LIVE-RC-005", "Live", "riefkohllaw.com/act-60-resource-center", "Chapter 3",
     "Chapter 3 of Act 60 offers a 4% corporate tax rate on eligible export services income.",
     "Tax Rate", "Critical", "PR", "EN", "Yes", "", "", ""],
    ["LIVE-RC-006", "Live", "riefkohllaw.com/act-60-resource-center", "Year-of-Move",
     "Special rules apply in the year you relocate. Learn the safe harbor under Treas. Reg. §1.937-1(f), its three conditions.",
     "Statutory Citation", "Critical", "Federal", "EN", "No", "", "", ""],
    ["LIVE-RC-007", "Live", "riefkohllaw.com/act-60-resource-center", "Estate Planning",
     "Puerto Rico residents born in a U.S. territory are classified as non-residents not citizens (NRNC) for estate and gift tax.",
     "Tax Rate", "Critical", "Federal", "EN", "No", "", "", ""],

    # ---- SPANISH PAGES ----
    ["LIVE-ES-001", "Live", "riefkohllaw.com/espanol", "Hero",
     "El Lcdo. Hans Riefkohl combina la experiencia de haber trabajado en DLA Piper, LUMA Energy y los tribunales federales.",
     "Professional Claim", "High", "General", "ES", "No", "", "", ""],
    ["LIVE-ES-002", "Live", "riefkohllaw.com/espanol", "Forced Heirship",
     "El Código Civil impone la legítima — una porción forzosa que limita cómo se distribuye su patrimonio.",
     "Legal Requirement", "Critical", "PR", "ES", "No", "", "", ""],
    ["LIVE-ES-003", "Live", "riefkohllaw.com/espanol", "Act 60 Compliance",
     "Hacienda y la Oficina de Incentivos Contributivos exigen cumplimiento continuo: informes anuales, donaciones, presencia física y más. Un error puede resultar en la revocación de su decreto.",
     "Legal Requirement", "Critical", "PR", "ES", "No", "", "", ""],
    ["LIVE-ES-004", "Live", "riefkohllaw.com/espanol", "Why Riefkohl",
     "El Lcdo. Riefkohl fue asociado en DLA Piper, oficial jurídico en el Tribunal Supremo de PR y en el Tribunal Federal, y gestionó transacciones de más de $100 millones en LUMA Energy.",
     "Professional Claim", "Critical", "General", "ES", "No", "", "", ""],
    ["LIVE-ES-005", "Live", "riefkohllaw.com/espanol-servicios", "Trust Pricing",
     "Nivel 1 — Fideicomiso sencillo de un beneficiario ($7,500 – $12,000). Nivel 2 — Fideicomiso de protección familiar ($12,000 – $18,000). Nivel 3 — Planificación sucesoral integral ($18,000 – $30,000). Nivel 4 — Planificación institucional ($30,000 – $50,000+).",
     "Financial Claim", "High", "General", "ES", "Yes", "", "", ""],
    ["LIVE-ES-006", "Live", "riefkohllaw.com/espanol-servicios", "Trust Law Cite",
     "Fideicomisos irrevocables bajo la Ley 219, le ayudamos a planificar con claridad — tomando en cuenta la legítima, la sociedad legal de gananciales y las particularidades del derecho civil puertorriqueño.",
     "Statutory Citation", "High", "PR", "ES", "No", "", "", ""],
    ["LIVE-ES-007", "Live", "riefkohllaw.com/espanol-servicios", "Bankruptcy",
     "Representamos acreedores en procesos de quiebra bajo el Título 11 del Código federal.",
     "Statutory Citation", "High", "Federal", "ES", "No", "", "", ""],
    ["LIVE-ES-008", "Live", "riefkohllaw.com/espanol-servicios", "Business Formation",
     "Organización de LLC: Desde $1,500. Constitución de corporación: Desde $2,000. Acuerdos de sociedad: Desde $2,500. Paquete de gobernanza corporativa: Desde $3,500.",
     "Financial Claim", "High", "General", "ES", "Yes", "", "", ""],
    ["LIVE-ES-009", "Live", "riefkohllaw.com/espanol-servicios", "Transactional",
     "Redacción y revisión de contratos: Desde $750. Contratos gubernamentales: Desde $3,000. Servicios de Ley 60: Desde $5,000.",
     "Financial Claim", "High", "General", "ES", "Yes", "", "", ""],

    # ---- HOMEPAGE ----
    ["LIVE-HP-001", "Live", "riefkohllaw.com/home-project", "Hero",
     "Backed by experience at DLA Piper, LUMA Energy, and the federal courts.",
     "Professional Claim", "High", "General", "EN", "No", "", "", ""],
    ["LIVE-HP-002", "Live", "riefkohllaw.com/home-project", "Differentiator",
     "From $100M+ transactions to small business formations, our counsel is informed by real-world experience at every scale.",
     "Professional Claim", "High", "General", "EN", "No", "", "", ""],
]

row_num = 2
for stmt in live_statements:
    add_row(ws1, row_num, stmt)
    row_num += 1

# ============================================================
# SHEET 2: TIME-SENSITIVE CLAIMS (filtered view)
# ============================================================
ws2 = setup_sheet(wb.create_sheet(), "Time-Sensitive Claims")
row_num = 2
for stmt in live_statements:
    if stmt[9] == "Yes":
        add_row(ws2, row_num, stmt)
        row_num += 1

# ============================================================
# SHEET 3: CRITICAL SEVERITY (filtered view)
# ============================================================
ws3 = setup_sheet(wb.create_sheet(), "Critical Severity")
row_num = 2
for stmt in live_statements:
    if stmt[6] == "Critical":
        add_row(ws3, row_num, stmt)
        row_num += 1

# ============================================================
# SHEET 4: PROFESSIONAL CLAIMS (About page, bio, credentials)
# ============================================================
ws4 = setup_sheet(wb.create_sheet(), "Professional Claims")
row_num = 2
for stmt in live_statements:
    if stmt[5] == "Professional Claim":
        add_row(ws4, row_num, stmt)
        row_num += 1

# ============================================================
# SHEET 5: FINANCIAL CLAIMS (pricing, costs, fees)
# ============================================================
ws5 = setup_sheet(wb.create_sheet(), "Financial Claims")
row_num = 2
for stmt in live_statements:
    if stmt[5] in ("Financial Claim", "Tax Rate"):
        add_row(ws5, row_num, stmt)
        row_num += 1

# ============================================================
# SHEET 6: SUMMARY DASHBOARD
# ============================================================
ws6 = wb.create_sheet("Summary Dashboard")
ws6.column_dimensions['A'].width = 30
ws6.column_dimensions['B'].width = 15
ws6.column_dimensions['C'].width = 15
ws6.column_dimensions['D'].width = 15

# Title
ws6.cell(row=1, column=1, value="Riefkohl Law — Liability Statement Database").font = Font(bold=True, size=14)
ws6.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True)
ws6.cell(row=3, column=1, value=f"Total Statements (Live Site): {len(live_statements)}").font = Font(bold=True, size=12)

# By Severity
ws6.cell(row=5, column=1, value="BY SEVERITY").font = Font(bold=True, size=11)
for header_col, header_val in enumerate(["Severity", "Count", "% of Total"], 1):
    c = ws6.cell(row=6, column=header_col, value=header_val)
    c.font = header_font; c.fill = header_fill
severities = ["Critical", "High", "Medium", "Low"]
for i, sev in enumerate(severities, 7):
    count = sum(1 for s in live_statements if s[6] == sev)
    ws6.cell(row=i, column=1, value=sev)
    ws6.cell(row=i, column=2, value=count)
    ws6.cell(row=i, column=3, value=f"{count/len(live_statements)*100:.1f}%")
    if sev == "Critical": ws6.cell(row=i, column=1).fill = critical_fill
    elif sev == "High": ws6.cell(row=i, column=1).fill = high_fill
    elif sev == "Medium": ws6.cell(row=i, column=1).fill = medium_fill
    elif sev == "Low": ws6.cell(row=i, column=1).fill = low_fill

# By Category
ws6.cell(row=13, column=1, value="BY CATEGORY").font = Font(bold=True, size=11)
for header_col, header_val in enumerate(["Category", "Count", "% of Total"], 1):
    c = ws6.cell(row=14, column=header_col, value=header_val)
    c.font = header_font; c.fill = header_fill
categories = {}
for s in live_statements:
    categories[s[5]] = categories.get(s[5], 0) + 1
for i, (cat, count) in enumerate(sorted(categories.items(), key=lambda x: -x[1]), 15):
    ws6.cell(row=i, column=1, value=cat)
    ws6.cell(row=i, column=2, value=count)
    ws6.cell(row=i, column=3, value=f"{count/len(live_statements)*100:.1f}%")

# By Page
cat_end = 15 + len(categories) + 1
ws6.cell(row=cat_end, column=1, value="BY SOURCE PAGE").font = Font(bold=True, size=11)
for header_col, header_val in enumerate(["Page", "Count", "Critical"], 1):
    c = ws6.cell(row=cat_end+1, column=header_col, value=header_val)
    c.font = header_font; c.fill = header_fill
pages = {}
page_critical = {}
for s in live_statements:
    page = s[2].split("/")[-1] if "/" in s[2] else s[2]
    pages[page] = pages.get(page, 0) + 1
    if s[6] == "Critical":
        page_critical[page] = page_critical.get(page, 0) + 1
for i, (page, count) in enumerate(sorted(pages.items(), key=lambda x: -x[1]), cat_end+2):
    ws6.cell(row=i, column=1, value=page)
    ws6.cell(row=i, column=2, value=count)
    ws6.cell(row=i, column=3, value=page_critical.get(page, 0))

# Time-sensitive count
ts_count = sum(1 for s in live_statements if s[9] == "Yes")
ts_row = cat_end + 2 + len(pages) + 1
ws6.cell(row=ts_row, column=1, value="TIME-SENSITIVE CLAIMS").font = Font(bold=True, size=11)
ws6.cell(row=ts_row+1, column=1, value=f"{ts_count} statements flagged as time-sensitive ({ts_count/len(live_statements)*100:.1f}% of total)")

# ============================================================
# SAVE
# ============================================================
output_path = "/Users/hansriefkohl/Downloads/Claude Code/Riefkohl Law updates/liability-statement-database.xlsx"
wb.save(output_path)
print(f"Database saved to: {output_path}")
print(f"Total live site statements: {len(live_statements)}")
print(f"Critical: {sum(1 for s in live_statements if s[6] == 'Critical')}")
print(f"High: {sum(1 for s in live_statements if s[6] == 'High')}")
print(f"Medium: {sum(1 for s in live_statements if s[6] == 'Medium')}")
print(f"Low: {sum(1 for s in live_statements if s[6] == 'Low')}")
print(f"Time-sensitive: {ts_count}")
